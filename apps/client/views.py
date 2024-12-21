from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import redirect, render
from django.urls import reverse
from openpyxl import load_workbook

import logging

logger = logging.getLogger(__name__)

from apps.users.decorators import (
    admin_or_manager_or_staff_required,
    admin_or_manager_required,
    admin_required,
)

from .forms import ClientForm, ImportClientsForm, SevenHillsRegistrationForm
from .models import Client, SevenHillsRegistration


# =================================== Fetch and display all clients details ===================================
@login_required
@admin_or_manager_or_staff_required
def client_list(request):
    queryset = Client.objects.all().order_by("id")

    search_query = request.GET.get("search")
    if search_query:
        queryset = queryset.filter(full_name__icontains=search_query)

    paginator = Paginator(queryset, 20)  # Show 20 records per page
    page = request.GET.get("page")

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        records = paginator.page(paginator.num_pages)

    return render(
        request,
        "sdms/client/client_list.html",
        {"records": records, "table_title": "Clients List"},
    )


# =================================== Register Client  ===================================


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def register_client(request):
    if request.method == "POST":
        form = ClientForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.success(
                request, "Record saved successfully!", extra_tags="bg-success"
            )
            return redirect("register_client")
        else:
            # Display an error message if the form is not valid
            messages.error(
                request,
                "There was an error saving the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )

    else:
        form = ClientForm()

    return render(
        request,
        "sdms/client/client_register.html",
        {"form_name": "Client Registration", "form": form},
    )


# =================================== Update client data ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def update_client(request, pk, template_name="sdms/client/client_update.html"):
    try:
        client_record = Client.objects.get(pk=pk)
    except Client.DoesNotExist:
        messages.error(request, "Client record not found!", extra_tags="bg-danger")
        return redirect("client_list")  # Or a relevant error page

    if request.method == "POST":
        form = ClientForm(request.POST, request.FILES, instance=client_record)
        if form.is_valid():
            form.save()

            messages.success(
                request, "Client record updated successfully!", extra_tags="bg-success"
            )
            return redirect("client_list")
    else:
        # Pre-populate the form with existing data
        form = ClientForm(instance=client_record)

    context = {"form_name": "Client Registration", "form": form}
    return render(request, template_name, context)


# =================================== Delete selected client ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_client(request, pk):
    records = Client.objects.get(id=pk)
    records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("client_list"))


# =================================== Process and Import Excel data ===================================
@login_required
@admin_required
@transaction.atomic
def import_client_data(request):
    if request.method == "POST":
        form = ImportClientsForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES.get("excel_file")
            if excel_file and excel_file.name.endswith(".xlsx"):
                try:
                    # Call process_and_import_data function
                    errors = process_and_import_data(excel_file)
                    if errors:
                        for error in errors:
                            messages.error(request, error, extra_tags="bg-danger")
                    else:
                        messages.success(
                            request,
                            "Data imported successfully!",
                            extra_tags="bg-success",
                        )
                except Exception as e:
                    messages.error(
                        request, f"Error importing data: {e}", extra_tags="bg-danger"
                    )
                return redirect("client_list")
            else:
                messages.error(
                    request, "Please upload a valid Excel file.", extra_tags="bg-danger"
                )
    else:
        form = ImportClientsForm()
    return render(
        request,
        "sdms/client/bulk_import.html",
        {"form_name": "Import Clients - Excel", "form": form},
    )


# Function to import Excel data
def process_and_import_data(excel_file):
    errors = []
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            fname = row[0].value
            picture = row[1].value
            reg_number = row[2].value
            mobile_telephone = row[3].value
            if fname is not None:
                try:
                    Client.objects.create(
                        full_name=fname,
                        picture=picture,
                        reg_number=reg_number,
                        mobile_telephone=mobile_telephone,
                    )
                except Exception as e:
                    errors.append(f"Error on row {row_num}: {e}")
            else:
                errors.append(f"Missing full name on row {row_num}")
    except Exception as e:
        errors.append(f"Failed to process the Excel file: {e}")
    return errors


# =================================== Delete all records at once ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_confirm(request):
    if request.method == "POST":
        Client.objects.all().delete()
        messages.info(request, "All records deleted!", extra_tags="bg-danger")
        return HttpResponseRedirect(reverse("client_list"))


# =================================== seven_hills registration ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def seven_hills_registration_view(request):
    if request.method == "POST":
        form = SevenHillsRegistrationForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.success(
                request, "Record saved successfully!", extra_tags="bg-success"
            )
            return redirect("seven_hills_registration")
        else:
            # Display error messages if the form is invalid
            messages.error(
                request,
                "There was an error saving the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )
    else:
        form = SevenHillsRegistrationForm()

    context = {
        "form_name": "Seven Hills Registration",
        "form": form,
    }

    return render(request, "sdms/client/seven_hills_register.html", context)


# =================================== Fetch and display all Seven Hills Registration details ===================================
@login_required
@admin_or_manager_or_staff_required
def seven_hills_list(request):
    # Fetch all records
    queryset = SevenHillsRegistration.objects.all().order_by("id")

    # Apply search filter
    search_query = request.GET.get("search")
    if search_query:
        queryset = queryset.filter(full_name__icontains=search_query)
        if not queryset.exists():
            messages.info(request, "No results found for your search.")

    # Paginate the filtered queryset
    paginator = Paginator(queryset, 20)  # Show 20 records per page
    page = request.GET.get("page")

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        records = paginator.page(1)
    except EmptyPage:
        records = paginator.page(paginator.num_pages)

    # Pass both the full queryset and paginated records to the template
    return render(
        request,
        "sdms/client/seven_hills_list.html",
        {
            "records": records,  # Paginated records for display
            "table_title": "Seven Hills Members List",
            "queryset": queryset,  # Full queryset if needed elsewhere
        },
    )


# =================================== Update Seven Hills data ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def update_seven_hills(
    request, pk, template_name="sdms/client/seven_hills_update.html"
):
    try:
        record = SevenHillsRegistration.objects.get(pk=pk)
    except SevenHillsRegistration.DoesNotExist:
        messages.error(request, "Record not found!", extra_tags="bg-danger")
        return redirect("seven_hills_list")  # Or a relevant error page

    if request.method == "POST":
        form = SevenHillsRegistrationForm(request.POST, request.FILES, instance=record)
        if form.is_valid():
            form.save()

            messages.success(
                request, "Record updated successfully!", extra_tags="bg-success"
            )
            return redirect("seven_hills_list")
    else:
        # Pre-populate the form with existing data
        form = SevenHillsRegistrationForm(instance=record)

    context = {"form_name": "Seven Hills Update", "form": form}
    return render(request, template_name, context)


# =================================== Delete selected Seven Hills ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_seven_hills(request, pk):
    records = SevenHillsRegistration.objects.get(id=pk)
    records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("seven_hills_list"))


# =================================== Fetch and display selected member details ===================================
@login_required
@admin_or_manager_or_staff_required
def seven_hills_details(request, pk):
    record = SevenHillsRegistration.objects.get(pk=pk)
    age = record.calculate_age()

    context = {"table_title": "Profile Report", "record": record, "age": age}
    return render(request, "sdms/client/seven_hills_profile_rpt.html", context)
