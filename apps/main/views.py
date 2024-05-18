# from formtools.wizard.views import SessionWizardView
import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import load_workbook

from apps.users.models import Contact

from .forms import (
    ChildCorrespondenceForm,
    ChildForm,
    ChildIncidentForm,
    ChildProfilePictureForm,
    ChildProgressForm,
    UploadForm,
    ChildDepartForm,
)
from .models import (
    Child, 
    ChildCorrespondence, 
    ChildIncident,
    ChildProfilePicture, 
    ChildProgress, 
    ChildDepart

)
# The getLogger() function is used to get a logger instance
logger = logging.getLogger(__name__)
# logger.info("Child ID received: %s", child_id)  # Log the child_id value


def home(request):
    return render(request, "users/home.html")


# =================================== The dashboard ===================================
@login_required
def dashboard(request):
    c_records = Child.objects.all().filter(is_departed="No")
    total_records = c_records.count()

    context = {
        "kids_registered": total_records,
        "total_no_of_kids": total_records,
    }
    return render(request, "main/dashboard.html", context)


# =================================== Fetch and display all children details ===================================
def child_list(request):
    # queryset = Child.objects.all().filter(is_departed="No").order_by("id").select_related("profile_picture")
    queryset = Child.objects.all().filter(is_departed="No").order_by("id")

    search_query = request.GET.get("search")
    if search_query:
        queryset = queryset.filter(full_name__icontains=search_query)

    paginator = Paginator(queryset, 25)  # Show 25 records per page
    page = request.GET.get("page")

    try:
        c_records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        c_records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        c_records = paginator.page(paginator.num_pages)

    return render(
        request,
        "main/child/manage_child.html",
        {"c_records": c_records, "table_title": "Children MasterList"},
    )


# =================================== Fetch and display selected child's details ===================================
@login_required
def child_details(request, pk):
    record = Child.objects.get(pk=pk)
    age = record.calculate_age()

    context = {"table_title": "Child Profile Report", "record": record, "age": age}
    return render(request, "main/child/child_profile_rpt.html", context)


# =================================== Register Child ===================================
@login_required
@transaction.atomic
def register_child(request):
    if request.method == "POST":
        form = ChildForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.info(
                request, "Record saved successfully!", extra_tags="bg-success"
            )

        else:
            # Display form errors
            return render(request, "main/child/register_child.html", {"form": form})
    else:
        form = ChildForm()
    return render(
        request,
        "main/child/register_child.html",
        {"form_name": "Child Registration", "form": form},
    )

# =================================== Update Child data ===================================
@login_required
@transaction.atomic
def update_child(request, pk, template_name="main/child/register_child.html"):
    try:
        child_record = Child.objects.get(pk=pk)
    except Child.DoesNotExist:
        messages.error(request, "Child record not found!")
        return redirect("child_list")  # Or a relevant error page

    if request.method == "POST":
        form = ChildForm(request.POST, request.FILES, instance=child_record)
        if form.is_valid():
            form.save()

            messages.success(request, "Child record updated successfully!")
            return redirect("child_list")  # Replace with the appropriate redirect URL
    else:
        # Pre-populate the form with existing data
        form = ChildForm(instance=child_record)

    context = {"form_name": "Child Registration", "form": form}
    return render(request, template_name, context)


# =================================== Deleted selected child ===================================
@login_required
@transaction.atomic
def delete_child(request, pk):
    c_records = Child.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("child_list"))


# =================================== Upload Profile Picture ===================================

@login_required
@transaction.atomic
def update_picture(request):
    if request.method == "POST":
        form = ChildProfilePictureForm(request.POST, request.FILES)
        if form.is_valid():
            child_id = request.POST.get("id")
            try:
                # Attempt to retrieve the child profile
                child_profile = Child.objects.get(id=child_id)
            except Child.DoesNotExist:
                # Handle the case where the child doesn't exist
                messages.error(request, "Child profile not found.")
                return redirect("update_picture")

            # Create a ChildProfilePicture instance
            new_picture = form.save(commit=False)
            new_picture.child = child_profile
            new_picture.is_current = True
            new_picture.save()

            # Update Child
            child_profile.picture = new_picture.picture
            child_profile.save()
            messages.success(request, "Profile picture updated successfully!")
            return redirect("update_picture")
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = ChildProfilePictureForm()

    # Retrieve all child objects
    children = Child.objects.all().filter(is_departed="No").order_by("id")

    return render(
        request,
        "main/child/profile_picture.html",
        {
            "form": form,
            "form_name": "Upload Child Profile Picture",
            "children": children,
        },
    )

# =================================== View Profile Pictures ===================================
@login_required
def profile_pictures(request):
    if request.method == "POST":
        child_id = request.POST.get("id")
        if child_id:
            selected_child = get_object_or_404(Child, id=child_id)
            profile_picture = ChildProfilePicture.objects.filter(child_id=child_id)
            children = Child.objects.all().filter(is_departed="No").order_by("id")
            return render(request, 'main/child/profile_picture_list.html', 
                          {"table_title": "Profile Pictures", "children": children, 
                           "child_name": selected_child.full_name, "prefix_id":selected_child.prefixed_id, 
                           'profile_picture': profile_picture})
        else:
            messages.error(request, "No child selected.")
    else:
        # Handle the GET request, show the form without results
        children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(request, 'main/child/profile_picture_list.html', 
                    {"table_title": "Profile Pictures", "children": children})

# =================================== Delete Profile Pictures ===================================
@login_required
@transaction.atomic
def delete_profile_picture(request, pk):
    c_records = ChildProfilePicture.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("profile_pictures"))

# =================================== Update Child Progress ===================================
@login_required
@transaction.atomic
def child_progress(request):
    if request.method == "POST":
        form = ChildProgressForm(request.POST)
        if form.is_valid():
            child_id = request.POST.get("id")
            child_instance = get_object_or_404(Child, pk=child_id)

            # Always create a new progress record explicitly
            child_progress = ChildProgress.objects.create(child=child_instance)

            # Populate progress data
            child_progress.name_of_school = form.cleaned_data["name_of_school"]
            child_progress.previous_schools = form.cleaned_data["previous_schools"]
            child_progress.education_level = form.cleaned_data["education_level"]
            child_progress.child_class = form.cleaned_data["child_class"]
            child_progress.best_subject = form.cleaned_data["best_subject"]
            child_progress.score = form.cleaned_data["score"]
            child_progress.co_curricular_activity = form.cleaned_data["co_curricular_activity"]
            child_progress.responsibility_at_school = form.cleaned_data["responsibility_at_school"]
            child_progress.future_plans = form.cleaned_data["future_plans"]
            child_progress.responsibility_at_home = form.cleaned_data["responsibility_at_home"]
            child_progress.notes = form.cleaned_data["notes"]
            child_progress.save()

            messages.success(request, "Child progress recorded successfully!")
            return redirect("child_progress")
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = ChildProgressForm()

    children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(
        request,
        "main/child/child_progress.html",
        {"form": form, "form_name": "Child Progress Form", "children": children},
    )

# =================================== View Child Progress Report ===================================
@login_required
def child_progress_report(request):
    if request.method == "POST":
        child_id = request.POST.get("id")
        if child_id:
            selected_child = get_object_or_404(Child, id=child_id)
            child_progress = ChildProgress.objects.filter(child_id=child_id)
            children = Child.objects.all().filter(is_departed="No").order_by("id")
            return render(request, 'main/child/progress_rpt.html', 
                          {"table_title": "Progress Report", "children": children, 
                           "child_name": selected_child.full_name, "prefix_id":selected_child.prefixed_id, 
                           'child_progress': child_progress})
        else:
            messages.error(request, "No child selected.")
    else:
        # Handle the GET request, show the form without results
        children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(request, 'main/child/progress_rpt.html', 
                    {"table_title": "Progress Report", "children": children})


 # =================================== Delete Progress Data ===================================   

@login_required
@transaction.atomic
def delete_progress(request, pk):
    c_records = ChildProgress.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("child_progress_report"))


# =================================== Update Child Correspondence ===================================

@login_required
@transaction.atomic
def child_correspondence(request):
    if request.method == "POST":
        form = ChildCorrespondenceForm(request.POST, request.FILES)
        if form.is_valid():
            child_id = request.POST.get("id")
            child_instance = get_object_or_404(Child, pk=child_id)

            # Always create a new correspondence record explicitly
            child_correspondence = ChildCorrespondence.objects.create(child=child_instance)

            # Populate correspondence data
            child_correspondence.correspondence_type = form.cleaned_data["correspondence_type"]
            child_correspondence.source = form.cleaned_data["source"]
            child_correspondence.attachment = form.cleaned_data["attachment"]
            child_correspondence.comment = form.cleaned_data["comment"]
            child_correspondence.save()

            messages.success(request, "Child correspondence recorded successfully!")
            return redirect("child_correspondence")
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = ChildCorrespondenceForm()

    children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(
        request,
        "main/child/child_correspondence.html",
        {"form": form, "form_name": "Child Correspondence Form", "children": children},
    )

# =================================== View Child Correspondence Report ===================================
@login_required
def child_correspondence_report(request):
    if request.method == "POST":
        child_id = request.POST.get("id")
        if child_id:
            selected_child = get_object_or_404(Child, id=child_id)
            child_correspondence = ChildCorrespondence.objects.filter(child_id=child_id)
            children = Child.objects.all().filter(is_departed="No").order_by("id")
            return render(request, 'main/child/correspondence_rpt.html', 
                          {"table_title": "Correspondence Report", "children": children, 
                           "child_name": selected_child.full_name, "prefix_id":selected_child.prefixed_id, 
                           'child_correspondence': child_correspondence})
        else:
            messages.error(request, "No child selected.")
    else:
        # Handle the GET request, show the form without results
        children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(request, 'main/child/correspondence_rpt.html', 
                    {"table_title": "Correspondence Report", "children": children})


# =================================== Delete Correspondence Data ===================================

@login_required
@transaction.atomic
def delete_correspondence(request, pk):
    c_records = ChildCorrespondence.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("child_correspondence_report"))


# =================================== Update Child Incident ===================================
@login_required
@transaction.atomic
def child_incident(request):
    if request.method == "POST":
        form = ChildIncidentForm(request.POST, request.FILES)
        if form.is_valid():
            child_id = request.POST.get("id")
            child_instance = get_object_or_404(Child, pk=child_id)

            # Always create a new incidence record explicitly
            child_incident = ChildIncident.objects.create(child=child_instance)

            # Populate incidence data
            child_incident.incident_date = form.cleaned_data["incident_date"]
            child_incident.description = form.cleaned_data["description"]
            child_incident.action_taken = form.cleaned_data["action_taken"]
            child_incident.results = form.cleaned_data["results"]
            child_incident.reported_by = form.cleaned_data["reported_by"]
            child_incident.followed_up_by = form.cleaned_data["followed_up_by"]
            child_incident.attachment = form.cleaned_data["attachment"]
            child_incident.save()

            messages.success(request, "Child incident recorded successfully!")
            return redirect("child_incident")
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = ChildIncidentForm()

    children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(
        request,
        "main/child/child_incident.html",
        {"form": form, "form_name": "Child Incident Form", "children": children},
    )


# =================================== Update Child Incident ===================================
@login_required
def child_incident_report(request):
    if request.method == "POST":
        child_id = request.POST.get("id")
        if child_id:
            selected_child = get_object_or_404(Child, id=child_id)
            child_incident = ChildIncident.objects.filter(child_id=child_id)
            children = Child.objects.all().filter(is_departed="No").order_by("id")
            return render(request, 'main/child/child_incident_rpt.html', 
                          {"table_title": "Incident Report", "children": children, 
                           "child_name": selected_child.full_name, "prefix_id":selected_child.prefixed_id, 
                           'child_incident': child_incident})
        else:
            messages.error(request, "No child selected.")
    else:
        # Handle the GET request, show the form without results
        children = Child.objects.all().filter(is_departed="No").order_by("id")
    return render(request, 'main/child/child_incident_rpt.html', 
                    {"table_title": "Incident Report", "children": children})

# =================================== Delete Incident Data ===================================

@login_required
@transaction.atomic
def delete_incident(request, pk):
    c_records = ChildIncident.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("child_incident_report"))


# =================================== Display User Feedback ===================================
@login_required
@transaction.atomic
def user_feedback(request):
    feedback = Contact.objects.all()
    return render(
        request,
        "users/user_feedback.html",
        {"table_title": "User Feedback", "feedback": feedback},
    )

# =================================== Delete User Feedback ===================================
@login_required
@transaction.atomic
def delete_feedback(request, pk):
    feedback = Contact.objects.get(id=pk)
    feedback.delete()
    messages.info(request, "Record deleted!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("users-feedback"))


# =================================== Add Child Depature ===================================
@login_required
@transaction.atomic
def child_departure(request):
    if request.method == "POST":
        form = ChildDepartForm(request.POST, request.FILES)
        if form.is_valid():
            child_id = request.POST.get("id")
            child_instance = get_object_or_404(Child, pk=child_id)

             # Create a ChildDepart instance
            child_depart = ChildDepart.objects.create(child=child_instance)
            child_depart.depart_date = form.cleaned_data["depart_date"]
            child_depart.depart_reason = form.cleaned_data["depart_reason"]
            child_depart.save()

            # Update Child status to "departed"
            child_instance.is_departed = "Yes"
            child_instance.save()

            messages.success(request, "Child departed successfully!")
            return redirect("child_departure")
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = ChildDepartForm()

    children = Child.objects.filter(is_departed="No").order_by("id") 
    return render(
        request,
        "main/child/depart.html",
        {"form": form, "form_name": "Child Depature Form", "children": children},
    )

# =================================== Child Depature Report ===================================
def depature_list(request):
    queryset = Child.objects.all().filter(is_departed="Yes").order_by("id").prefetch_related("departures")

    search_query = request.GET.get("search")
    if search_query:
        queryset = queryset.filter(full_name__icontains=search_query)

    paginator = Paginator(queryset, 25)  # Show 10 records per page
    page = request.GET.get("page")

    try:
        c_records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        c_records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        c_records = paginator.page(paginator.num_pages)

    return render(
        request,
        "main/child/depature_list.html",
        {"c_records": c_records, "table_title": "Departed Children"},
    )

# =================================== Reinstate departed child ===================================
@login_required
@transaction.atomic
def reinstate_child(request, pk):
    child = get_object_or_404(Child, id=pk)
    
    if request.method == 'POST':
        child.is_departed = "No"
        child.save()
        messages.success(request, "Child reinstated successfully!")

        return redirect("depature_list")
    
    return render(request, 'main/child/depature_list.html', {'child': child})


# =================================== Process and Import Excel data ===================================
@login_required
@transaction.atomic
def import_data(request):
    if request.method == "POST":
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                # Call process_and_import_data function
                process_and_import_data(excel_file)
                messages.success(
                    request, "Data imported successfully!", extra_tags="bg-success"
                )
            except Exception as e:
                messages.error(
                    request, f"Error importing data: {e}", extra_tags="bg-danger"
                )  # Handle unexpected errors
            return redirect("imported_data")  # Replace with your redirect URL
    else:
        form = UploadForm()
    return render(
        request,
        "main/child/bulk_import.html",
        {"form_name": "Import Excel Data", "form": form},
    )


# Function to import Excel data
@login_required
@transaction.atomic
def process_and_import_data(excel_file):
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            fname = row[0].value
            preferred_name = row[1].value
            residence = row[2].value
            tribe = row[3].value
            gender = row[4].value
            date_of_birth = row[5].value
            weight = row[6].value
            height = row[7].value
            c_interest = row[8].value
            is_child_in_school = row[9].value
            is_sponsored = row[10].value  # True if
            father_name = row[11].value
            is_father_alive = row[12].value
            father_description = row[13].value
            mother_name = row[14].value
            is_mother_alive = row[15].value
            mother_description = row[16].value
            guardian = row[17].value
            guardian_contact = row[18].value
            relationship_with_guardian = row[19].value
            siblings = row[20].value
            background_info = row[21].value
            health_status = row[22].value
            responsibility = row[23].value
            relationship_with_christ = row[24].value
            religion = row[25].value
            prayer_request = row[26].value
            year_enrolled = row[27].value
            is_departed = row[28].value
            staff_comment = row[29].value
            compiled_by = row[30].value
            if fname is not None:
                obj = Child.objects.create(
                    full_name=fname,
                    preferred_name=preferred_name,
                    residence=residence,
                    tribe=tribe,
                    gender=gender,
                    date_of_birth=date_of_birth,
                    weight=weight,
                    height=height,
                    c_interest=c_interest,
                    is_child_in_school=is_child_in_school,
                    is_sponsored=is_sponsored,
                    father_name=father_name,
                    is_father_alive=is_father_alive,
                    father_description=father_description,
                    mother_name=mother_name,
                    is_mother_alive=is_mother_alive,
                    mother_description=mother_description,
                    guardian=guardian,
                    guardian_contact=guardian_contact,
                    relationship_with_guardian=relationship_with_guardian,
                    siblings=siblings,
                    background_info=background_info,
                    health_status=health_status,
                    responsibility=responsibility,
                    relationship_with_christ=relationship_with_christ,
                    religion=religion,
                    prayer_request=prayer_request,
                    year_enrolled=year_enrolled,
                    is_departed=is_departed,
                    staff_comment=staff_comment,
                    compiled_by=compiled_by,
                )
                obj.save()
    except Exception as e:
        raise e  # Reraise the exception for better error handling at the view level


# =================================== Fetch and display imported data ===================================
@login_required
@transaction.atomic
def import_details(request):
    records = Child.objects.all().filter(is_departed="No")
    return render(
        request,
        "main/child/imported_data.html",
        {"table_title": "Imported Excel Data", "records": records},
    )


# =================================== Delete selected individual ===================================
@login_required
@transaction.atomic
def delete_excel_data(request, pk):
    record_imported = Child.objects.get(id=pk)
    record_imported.delete()
    messages.info(request, "Record deleted!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("imported_data"))


# =================================== Delete all records at once ===================================
@login_required
@transaction.atomic
def delete_confirmation(request):
    if request.method == "POST":
        Child.objects.all().delete()
        messages.info(request, "All records deleted!", extra_tags="bg-danger")
        return HttpResponseRedirect(reverse("imported_data"))


# =================================== More ===================================
