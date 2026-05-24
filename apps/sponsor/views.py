import logging

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.management import call_command
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponseBadRequest, HttpResponseRedirect
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from openpyxl import load_workbook

from apps.users.decorators import (
    admin_or_manager_or_staff_required,
    admin_or_manager_required,
    admin_required,
)
from apps.finance.services import (
    empty_sponsor_portal_payment_summary,
    get_sponsor_program_payment_report_context,
    get_sponsor_portal_payment_summary,
)

from .forms import (
    DonorForm,
    SponsorDepartForm,
    SponsorFeedbackForm,
    SponsorForm,
    SponsorUploadForm,
)
from .models import (
    Donor,
    Sponsor,
    SponsorDeparture,
    SponsorFeedback,
    sponsorship_type_flags,
)
from .services import send_sponsor_feedback_email

logger = logging.getLogger(__name__)


def _get_request_sponsor(request):
    profile = getattr(request.user, "profile", None)
    sponsor = getattr(profile, "sponsor", None)

    if sponsor is None and request.user.email:
        sponsor = (
            Sponsor.objects.active_real_supporters()
            .filter(email__iexact=request.user.email)
            .first()
        )

    return sponsor


def _sponsor_support_records(sponsor):
    from apps.sponsorship.models import ChildSponsorship, StaffSponsorship

    child_sponsorships = (
        ChildSponsorship.objects.select_related("child", "sponsor")
        .filter(sponsor=sponsor)
        .order_by("-is_active", "child__full_name")
    )
    staff_sponsorships = (
        StaffSponsorship.objects.select_related("staff", "sponsor")
        .filter(sponsor=sponsor)
        .order_by("-is_active", "staff__first_name", "staff__last_name")
    )
    return child_sponsorships, staff_sponsorships


@login_required
def sponsor_portal(request):
    sponsor = _get_request_sponsor(request)

    child_sponsorships = []
    staff_sponsorships = []
    active_child_count = 0
    active_staff_count = 0
    expected_amount = 0
    payment_summary = empty_sponsor_portal_payment_summary()
    feedback_form = SponsorFeedbackForm()
    if sponsor is not None:
        child_sponsorships, staff_sponsorships = _sponsor_support_records(sponsor)
        active_child_count = child_sponsorships.filter(is_active=True).count()
        active_staff_count = staff_sponsorships.filter(is_active=True).count()
        expected_amount = sponsor.expected_amt
        payment_summary = get_sponsor_portal_payment_summary(sponsor)

    total_active_count = active_child_count + active_staff_count

    return render(
        request,
        "sponsor/sponsor_portal.html",
        {
            "sponsor": sponsor,
            "child_sponsorships": child_sponsorships,
            "staff_sponsorships": staff_sponsorships,
            "active_child_count": active_child_count,
            "active_staff_count": active_staff_count,
            "active_count": total_active_count,
            "expected_amount": expected_amount,
            "feedback_form": feedback_form,
            **payment_summary,
        },
    )


@login_required
@transaction.atomic
def submit_sponsor_feedback(request):
    sponsor = _get_request_sponsor(request)
    if sponsor is None:
        messages.error(
            request,
            "Your user account is not linked to a sponsor profile yet.",
            extra_tags="bg-danger",
        )
        return redirect("sponsor_portal")

    if request.method != "POST":
        return HttpResponseBadRequest("Invalid request")

    form = SponsorFeedbackForm(request.POST)
    if not form.is_valid():
        messages.error(
            request,
            "Please check the feedback form and try again.",
            extra_tags="bg-danger",
        )
        return redirect("sponsor_portal")

    feedback = form.save(commit=False)
    feedback.sponsor = sponsor
    feedback.submitted_by = request.user
    feedback.save()
    transaction.on_commit(lambda: send_sponsor_feedback_email(feedback))

    messages.success(
        request,
        "Thank you. Your feedback has been sent to the sponsorship team.",
        extra_tags="bg-success",
    )
    return redirect("sponsor_portal")


@login_required
@admin_or_manager_or_staff_required
def sponsor_feedback_report(request):
    feedback_list = SponsorFeedback.objects.with_related()
    paginator = Paginator(feedback_list, 25)
    page = request.GET.get("page")

    try:
        feedback = paginator.page(page)
    except PageNotAnInteger:
        feedback = paginator.page(1)
    except EmptyPage:
        feedback = paginator.page(paginator.num_pages)

    return render(
        request,
        "sponsor/sponsor_feedback_report.html",
        {
            "table_title": "Sponsor Feedback",
            "feedback": feedback,
            "total_feedback": feedback_list.count(),
        },
    )


@login_required
@admin_or_manager_required
@transaction.atomic
def mark_sponsor_feedback_reviewed(request, feedback_id):
    feedback = get_object_or_404(SponsorFeedback, id=feedback_id)

    if request.method == "POST":
        feedback.status = SponsorFeedback.Status.REVIEWED
        feedback.save(update_fields=["status", "updated_at"])
        messages.success(request, "Sponsor feedback marked as reviewed.", extra_tags="bg-success")
        return redirect("sponsor_feedback_report")

    return HttpResponseBadRequest("Invalid request")


def _sponsor_payment_report_context(sponsor, payment_model, beneficiary_type):
    payments = payment_model.objects.none()
    total_amount = 0
    payment_count = 0
    latest_payment = None
    yearly_totals = []

    if sponsor is not None:
        payments = (
            payment_model.objects.select_related(beneficiary_type, "sponsor")
            .filter(sponsor=sponsor, is_valid=True)
            .order_by("-payment_date", "-id")
        )
        total_amount = payments.aggregate(total=Sum("amount"))["total"] or 0
        payment_count = payments.count()
        latest_payment = payments.first()
        yearly_totals = payments.values("payment_year").annotate(
            total=Sum("amount")
        ).order_by("-payment_year")

    return {
        "sponsor": sponsor,
        "payments": payments,
        "total_amount": total_amount,
        "payment_count": payment_count,
        "latest_payment": latest_payment,
        "yearly_totals": yearly_totals,
    }


@login_required
def sponsor_child_payment_report(request):
    from apps.finance.models import ChildPayments

    sponsor = _get_request_sponsor(request)
    context = _sponsor_payment_report_context(sponsor, ChildPayments, "child")
    context.update(
        {
            "report_title": "Child Sponsorship Payment Report",
            "report_subtitle": "Validated child sponsorship payments linked to your sponsor account.",
            "beneficiary_label": "Child",
            "report_kind": "child",
        }
    )
    return render(request, "sponsor/sponsor_payment_report.html", context)


@login_required
def sponsor_staff_payment_report(request):
    from apps.finance.models import StaffPayments

    sponsor = _get_request_sponsor(request)
    context = _sponsor_payment_report_context(sponsor, StaffPayments, "staff")
    context.update(
        {
            "report_title": "Staff Sponsorship Payment Report",
            "report_subtitle": "Validated staff sponsorship payments linked to your sponsor account.",
            "beneficiary_label": "Staff",
            "report_kind": "staff",
        }
    )
    return render(request, "sponsor/sponsor_payment_report.html", context)


@login_required
def sponsor_program_payment_report(request, program_group):
    sponsor = _get_request_sponsor(request)
    context = get_sponsor_program_payment_report_context(sponsor, program_group)
    return render(request, "sponsor/sponsor_payment_report.html", context)


# =================================== Sponsors List ===================================
@login_required
@admin_or_manager_or_staff_required
def sponsor_list(request):
    search_query = request.GET.get("search", "").strip()
    queryset = Sponsor.objects.active_real_supporters().order_by("id")

    if search_query:
        queryset = queryset.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(mobile_telephone__icontains=search_query)
            | Q(business_telephone__icontains=search_query)
        )

    paginator = Paginator(queryset, 50)
    page = request.GET.get("page")

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        records = paginator.page(1)
    except EmptyPage:
        records = paginator.page(paginator.num_pages)

    return render(
        request,
        "sponsor/sponsor_details.html",
        {
            "records": records,
            "table_title": "Sponsors List",
            "search_query": search_query,
            "total_sponsors": Sponsor.objects.active_real_supporters().count(),
            "child_sponsors": Sponsor.objects.active_real_supporters().filter(is_child_sponsor=True).count(),
            "staff_sponsors": Sponsor.objects.active_real_supporters().filter(is_staff_sponsor=True).count(),
            "family_supporters": Sponsor.objects.active_real_supporters().filter(is_family_supporter=True).count(),
        },
    )


# =================================== Register Sponsor ===================================


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def register_sponsor(request):
    if request.method == "POST":
        form = SponsorForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.info(
                request, "Sponsor added successfully!", extra_tags="bg-success"
            )
            return redirect("register_sponsor")
        else:
            # Display an error message if the form is not valid
            messages.error(
                request,
                "There was an error saving the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )
    else:
        form = SponsorForm()
    return render(
        request,
        "sponsor/sponsor_register.html",
        {"form_name": "Sponsor Registration", "form": form},
    )


# ===================================  Donor List ===================================
@login_required
def donor_list_view(request):
    search_query = request.GET.get("search", "")
    donors = Donor.objects.all().order_by("-created_at")

    if search_query:
        donors = donors.filter(full_name__icontains=search_query)

    # Pagination
    paginator = Paginator(donors, 50)
    page_number = request.GET.get("page")
    donors_page = paginator.get_page(page_number)

    return render(
        request,
        "sponsor/donor_list.html",
        {
            "donors": donors_page,
            "table_title": "List of Other Donors",
            "search_query": search_query,
        },
    )


# =================================== Register Donor ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def add_donor_view(request):
    if request.method == "POST":
        form = DonorForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.info(request, "Donor added successfully!", extra_tags="bg-success")
            return redirect("add_donor")
        else:
            # Display an error message if the form is not valid
            messages.error(
                request,
                "There was an error saving the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )
    else:
        form = DonorForm()
    return render(
        request,
        "sponsor/add_donor.html",
        {"form_name": "Register Other Donors", "form": form},
    )


# =================================== Update Donor ===================================
@login_required
def update_donor_view(request, donor_id):
    donor = get_object_or_404(Donor, id=donor_id)

    if request.method == "POST":
        form = DonorForm(request.POST, instance=donor)
        if form.is_valid():
            form.save()
            messages.success(
                request, "Donor updated successfully!", extra_tags="bg-success"
            )
            return redirect("donor_list")
        else:
            messages.error(request, "Form is invalid.", extra_tags="bg-danger")
    else:
        form = DonorForm(instance=donor)

    return render(
        request,
        "sponsor/update_donor.html",
        {"form": form, "form_name": "Update Donor"},
    )


# =================================== Delete Donor ===================================
@login_required
@admin_or_manager_required
def delete_donor_view(request, donor_id):
    donor = get_object_or_404(Donor, id=donor_id)

    # Automatically delete the donor without confirmation
    donor.delete()
    messages.success(request, "Donor deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("donor_list"))


# =================================== Update Sponsor data ===================================
@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def update_sponsor(request, pk, template_name="sponsor/sponsor_update.html"):
    try:
        sponsor_record = Sponsor.objects.get(pk=pk)
    except Sponsor.DoesNotExist:
        messages.error(request, "Record not found!", extra_tags="bg-danger")
        return redirect("sponsor_list")  # Or a relevant error page

    if request.method == "POST":
        form = SponsorForm(request.POST, request.FILES, instance=sponsor_record)
        if form.is_valid():
            form.save()

            messages.success(
                request, "Record updated successfully!", extra_tags="bg-success"
            )
            return redirect("sponsor_list")
        else:
            # Display an error message if the form is not valid
            messages.error(
                request,
                "There was an error updating the record. Please check the form for errors.",
                extra_tags="bg-danger",
            )
    else:
        # Pre-populate the form with existing data
        form = SponsorForm(instance=sponsor_record)

    context = {"form_name": "Sponsor Registration", "form": form}
    return render(request, template_name, context)


# =================================== Delete selected Sponsor ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_sponsor(request, pk):
    records = Sponsor.objects.get(id=pk)
    records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("sponsor_list"))


# =================================== Depart Sponsor ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def sponsor_departure(request):
    if request.method == "POST":
        form = SponsorDepartForm(request.POST, request.FILES)
        if form.is_valid():
            sponsor_id = request.POST.get("id")
            if not sponsor_id:
                messages.error(request, "Please select a sponsor.", extra_tags="bg-danger")
                return redirect("sponsor_departure")
            sponsor_instance = get_object_or_404(Sponsor, pk=sponsor_id)

            # Create a sponsorDepart instance
            sponsor_depart = SponsorDeparture.objects.create(sponsor=sponsor_instance)
            sponsor_depart.departure_date = form.cleaned_data["departure_date"]
            sponsor_depart.departure_reason = form.cleaned_data["departure_reason"]
            sponsor_depart.save()

            # Update sponsor status to "departed"
            sponsor_instance.is_departed = True
            sponsor_instance.save()

            messages.success(
                request, "Sponsor departed successfully!", extra_tags="bg-success"
            )
            return redirect("sponsor_departure")
        else:
            messages.error(request, "Form is invalid.", extra_tags="bg-danger")
    else:
        form = SponsorDepartForm()

    sponsors = Sponsor.objects.active_real_supporters().order_by("first_name", "last_name", "id")
    return render(
        request,
        "sponsor/sponsor_depature.html",
        {
            "form": form,
            "form_name": "Sponsor Departure Form",
            "sponsors": sponsors,
            "active_sponsor_count": sponsors.count(),
            "departed_sponsor_count": Sponsor.objects.departed_real_supporters().count(),
        },
    )


# =================================== sponsor Depature Report ===================================
@login_required
@admin_or_manager_or_staff_required
def sponsor_depature_list(request):
    queryset = (
        Sponsor.objects.departed_real_supporters()
        .order_by("id")
        .prefetch_related("departures")
    )

    search_query = request.GET.get("search", "").strip()
    if search_query:
        queryset = queryset.filter(
            Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(email__icontains=search_query)
            | Q(mobile_telephone__icontains=search_query)
        )

    paginator = Paginator(queryset, 50)
    page = request.GET.get("page")

    try:
        records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        records = paginator.page(paginator.num_pages)

    return render(
        request,
        "sponsor/sponsor_depature_list.html",
        {
            "records": records,
            "table_title": "Departed Sponsors",
            "search_query": search_query,
            "departed_sponsor_count": queryset.count(),
        },
    )


# =================================== Reinstate departed sponsor ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def reinstate_sponsor(request, pk):
    sponsor = get_object_or_404(Sponsor, id=pk)

    if request.method == "POST":
        sponsor.is_departed = False
        sponsor.save()
        messages.success(
            request, "Sponsor reinstated successfully!", extra_tags="bg-success"
        )

        return redirect("sponsor_depature_list")

    return render(request, "sponsor/sponsor_depature_list.html", {"sponsor": sponsor})


# =================================== Process and Import Excel data ===================================
@login_required
@admin_required
@transaction.atomic
def import_sponsor_data(request):
    if request.method == "POST":
        form = SponsorUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["excel_file"]
            try:
                # Call process_and_import_data function
                process_and_import_data(excel_file)
                messages.success(
                    request, "Data imported successfully!", extra_tags="bg-success"
                )
            except Exception as e:
                messages.error(
                    request, f"Error importing data: {e}", extra_tags="bg-danger"
                )  # Handle unexpected errors
            return redirect("import_sponsor_data")  # Replace with your redirect URL
    else:
        form = SponsorUploadForm()
    return render(
        request,
        "sponsor/import_sponsors.html",
        {"form_name": "Import Sponsors - Excel", "form": form},
    )


# Function to import Excel data
def parse_boolean(value):
    """Convert values like 'Yes' or 'No' to boolean."""
    if value in ["Yes", "yes", True]:
        return True
    elif value in ["No", "no", False]:
        return False
    return None


def process_and_import_data(excel_file):
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            data = {
                "first_name": row[0].value,
                "last_name": row[1].value,
                "gender": row[2].value,
                "email": row[3].value,
                "sponsorship_type": row[4].value,
                "expected_amt": row[5].value,
                "job_title": row[6].value,
                "region": row[7].value,
                "town": row[8].value,
                "origin": row[9].value,
                "business_telephone": row[10].value,
                "mobile_telephone": row[11].value,
                "city": row[12].value,
                "start_date": row[13].value,
                "first_street_address": row[14].value,
                "second_street_address": row[15].value,
                "zip_code": row[16].value,
                "is_departed": parse_boolean(row[17].value),
                "comment": row[18].value,
            }
            data.update(sponsorship_type_flags(data["sponsorship_type"]))

            # Validate and log data
            logger.debug(f"Processing data: {data}")

            # Ensure values conform to the constraints
            if not data["first_name"]:
                logger.warning(f"Skipping row with missing first_name: {data}")
                continue  # Skip rows with missing required fields

            # Save the record
            obj = Sponsor(**data)
            obj.save()
    except Exception as e:
        logger.error(
            f"Error importing data: {e}", exc_info=True
        )  # Log error with traceback
        raise e


# =================================== Fetch and display imported data ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def imported_sponsors(request):
    records = Sponsor.objects.active_real_supporters().order_by("id")
    return render(
        request,
        "sponsor/imported_sponsors_rpt.html",
        {"table_title": "Imported Sponsors - Excel", "records": records},
    )


# =================================== Delete all sponsors at once ===================================
@login_required
@admin_or_manager_required
@transaction.atomic
def delete_sponsors(request):
    if request.method == "POST":
        Sponsor.objects.all().delete()
        messages.info(request, "All records deleted!", extra_tags="bg-danger")
        return HttpResponseRedirect(reverse("imported_sponsors"))


# ===================================  'Update a prefix + sign' ===================================
@login_required
@admin_required
def update_sponsor_contacts(request):
    if request.method == "POST":
        # Call the management command and handle success or failure
        try:
            call_command(
                "sponsor_contacts"
            )  # Replace "sponsor_contacts" with your actual command name
            messages.success(
                request,
                "Sponsors contacts updated successfully!",
                extra_tags="bg-success",
            )
            logger.info("Successfully updated sponsors contacts.")
        except Exception as e:
            messages.error(
                request,
                f"Error updating sponsors contacts: {e}",
                extra_tags="bg-danger",
            )
            logger.error(f"Error updating sponsors contacts: {e}", exc_info=True)

        # Redirect to avoid re-posting data on refresh
        return HttpResponseRedirect(reverse("imported_sponsors"))

    # Render the form if not a POST request
    return render(request, "sponsor/imported_sponsors_rpt.html")
