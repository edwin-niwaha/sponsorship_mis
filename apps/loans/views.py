import logging
from datetime import date, datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal

import pytz
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import transaction
from django.db.models import Count, Q, Sum
from django.http import Http404, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.html import strip_tags
from openpyxl import load_workbook

from apps.client.models import Client
from apps.users.decorators import (
    admin_or_manager_or_staff_required,
    admin_or_manager_required,
    admin_required,
)

from .forms import (
    ChartOfAccountsForm,
    ClientSelfServiceLoanApplicationForm,
    ImportCOAForm,
    ImportLoansForm,
    LoanAllDisbursementForm,
    LoanApplicationDocumentForm,
    LoanApplicationForm,
    LoanApplicationUpdateForm,
    LoanDisbursementForm,
    LoanPenaltyForm,
    LoanRepaymentForm,
    LoanReportFilterForm,
    StaffLoanApplicationDocumentForm,
)
from .models import (
    ChartOfAccounts,
    Loan,
    LoanApplicationDocument,
    LoanPenalty,
    LoanRepayment,
    TransactionHistory,
)
from .services.reporting import (
    ReportColumn,
    export_rows_csv,
    filtered_loans,
    group_rows_by_bucket,
    loan_financial_row,
    paginate_rows,
    parse_report_filters,
    portfolio_at_risk_summary,
    repayment_rows,
    summarize_amounts,
)
from .tasks import (
    send_html_email_task,
    send_loan_application_email_task,
    send_loan_approval_notification_task,
)

logger = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────


def get_loan_queryset(search_query=None):
    """Base queryset used by every loan list view."""
    qs = (
        Loan.objects.select_related("borrower", "account")
        .prefetch_related("disbursements")
        .order_by("id")
    )
    if search_query:
        qs = qs.filter(borrower__full_name__icontains=search_query)
    return qs


def paginate_queryset(queryset, page_number, per_page=50):
    paginator = Paginator(queryset, per_page)
    try:
        return paginator.page(page_number)
    except PageNotAnInteger:
        return paginator.page(1)
    except EmptyPage:
        return paginator.page(paginator.num_pages)


def loan_queryset_report_summary(queryset):
    """Small report totals for loan list pages without relying on DB-only properties."""
    loans = list(queryset)
    total_principal = sum((loan.principal_amount or Decimal("0.00")) for loan in loans)
    total_interest = sum((loan.total_interest or Decimal("0.00")) for loan in loans)
    return {
        "loan_count": len(loans),
        "total_principal": total_principal,
        "total_interest": total_interest,
        "total_exposure": total_principal + total_interest,
    }


def _is_ajax(request):
    return request.headers.get("X-Requested-With") == "XMLHttpRequest"


def _json_or_message(request, success, message, status=200, redirect_url=None):
    """Unified response helper for views that support both AJAX and normal POST."""
    if _is_ajax(request):
        return JsonResponse({"success": success, "message": message}, status=status)
    tag = "bg-success" if success else "bg-danger"
    fn = messages.success if success else messages.error
    fn(request, message, extra_tags=tag)
    return None  # caller should redirect


def _reverse_penalty_balance(penalty, user):
    """Reverse the unpaid part of a penalty and keep the original audit trail."""
    amount = penalty.remaining_amount or Decimal("0.00")
    if amount <= 0:
        return Decimal("0.00")

    try:
        income_account = ChartOfAccounts.objects.get(account_number="5030")
    except ChartOfAccounts.DoesNotExist as exc:
        raise ValidationError(
            "Loan Interest Income account (5030) does not exist."
        ) from exc

    reversal_date = timezone.localdate()
    description = f"Penalty reversal for Loan {penalty.loan.id}: penalty #{penalty.id}"
    TransactionHistory.objects.create(
        loan=penalty.loan,
        transaction_date=reversal_date,
        amount=amount,
        transaction_type="credit",
        account=penalty.account,
        description=description,
    )
    TransactionHistory.objects.create(
        loan=penalty.loan,
        transaction_date=reversal_date,
        amount=amount,
        transaction_type="debit",
        account=income_account,
        description=description,
    )
    penalty.remaining_amount = Decimal("0.00")
    penalty.is_deleted = True
    penalty.deleted_at = timezone.now()
    penalty.deleted_by = user if getattr(user, "is_authenticated", False) else None
    penalty.save(
        update_fields=[
            "remaining_amount",
            "is_deleted",
            "deleted_at",
            "deleted_by",
            "updated_at",
        ]
    )
    return amount


def _get_self_service_client(user):
    profile = getattr(user, "profile", None)
    if profile and profile.client_id:
        return profile.client
    email = (getattr(user, "email", "") or "").strip()
    if not email:
        return None
    return (
        Client.objects.filter(email__iexact=email)
        .exclude(email="no-email@example.com")
        .first()
    )


def _self_service_loan_queryset(user):
    client = _get_self_service_client(user)
    if client is None:
        return Loan.objects.none(), None
    return (
        Loan.objects.filter(borrower=client)
        .select_related("borrower", "applied_by")
        .prefetch_related("documents", "repayments", "disbursements")
        .order_by("-created_at", "-id")
    ), client


# ─────────────────────────────────────────────────────────────────────────────
# Aging / overdue helper  (used by loan_aging_report & loan_arrears_report)
# ─────────────────────────────────────────────────────────────────────────────


def compute_installment_based_days_overdue(loan: Loan, today: date) -> dict:
    """
    Installment-based aging.  Uses loan.monthly_installment (model @property)
    so the calculation is always consistent with the model.
    """
    disbursement_date = loan.disbursement_date
    term_months = loan.loan_period_months

    _empty = {
        "days_overdue": 0,
        "next_due_date": None,
        "final_due_date": None,
        "monthly_installment": Decimal("0.00"),
        "total_repayable": Decimal("0.00"),
        "installments_due_by_now": 0,
        "total_due_by_today": Decimal("0.00"),
        "total_paid_pi": Decimal("0.00"),
        "shortfall": Decimal("0.00"),
        "first_unpaid_due_date": None,
    }

    if not disbursement_date or not term_months or term_months <= 0:
        if loan.due_date:
            days_overdue = max((today - loan.due_date).days, 0)
            return {
                **_empty,
                "days_overdue": days_overdue,
                "next_due_date": loan.due_date if days_overdue > 0 else None,
                "final_due_date": loan.due_date,
            }
        return _empty

    # Use model properties — single source of truth
    total_repayable = loan.total_repayable
    monthly_installment = loan.monthly_installment
    final_due_date = disbursement_date + relativedelta(months=term_months)

    paid = loan.repayments.filter(repayment_date__lte=today).aggregate(
        total_principal=Sum("principal_payment"),
        total_interest=Sum("interest_payment"),
    )
    total_paid_pi = (paid["total_principal"] or Decimal("0.00")) + (
        paid["total_interest"] or Decimal("0.00")
    )

    # Pass 1 — count all installments due up to today
    installments_due_by_now = sum(
        1
        for n in range(1, term_months + 1)
        if disbursement_date + relativedelta(months=n) <= today
    )

    if installments_due_by_now == 0:
        return {
            **_empty,
            "next_due_date": disbursement_date + relativedelta(months=1),
            "final_due_date": final_due_date,
            "monthly_installment": monthly_installment,
            "total_repayable": total_repayable,
            "total_paid_pi": total_paid_pi,
        }

    total_due_by_today = (monthly_installment * installments_due_by_now).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )

    # Pass 2 — find first installment not yet covered
    first_unpaid_due_date = None
    for n in range(1, installments_due_by_now + 1):
        cumulative_due = (monthly_installment * n).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
        if total_paid_pi + Decimal("0.01") < cumulative_due:
            first_unpaid_due_date = disbursement_date + relativedelta(months=n)
            break

    if first_unpaid_due_date is None:
        next_installment_no = installments_due_by_now + 1
        next_due_date = (
            disbursement_date + relativedelta(months=next_installment_no)
            if next_installment_no <= term_months
            else None
        )
        return {
            **_empty,
            "next_due_date": next_due_date,
            "final_due_date": final_due_date,
            "monthly_installment": monthly_installment,
            "total_repayable": total_repayable,
            "installments_due_by_now": installments_due_by_now,
            "total_due_by_today": total_due_by_today,
            "total_paid_pi": total_paid_pi,
        }

    shortfall = (total_due_by_today - total_paid_pi).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    days_overdue = max((today - first_unpaid_due_date).days, 0)

    return {
        "days_overdue": days_overdue,
        "next_due_date": first_unpaid_due_date,
        "final_due_date": final_due_date,
        "monthly_installment": monthly_installment,
        "total_repayable": total_repayable,
        "installments_due_by_now": installments_due_by_now,
        "total_due_by_today": total_due_by_today,
        "total_paid_pi": total_paid_pi,
        "shortfall": shortfall,
        "first_unpaid_due_date": first_unpaid_due_date,
    }


def _get_next_payment_date(loan: Loan) -> date | None:
    """
    Returns the due date of the next unpaid installment.
    Used by portfolio, non-performing-loans, and arrears report views.
    Reuses loan.monthly_installment so logic stays consistent with the model.
    """
    if not loan.disbursement_date or not loan.loan_period_months:
        return None
    total_paid = loan.repayments.aggregate(total=Sum("principal_payment"))[
        "total"
    ] or Decimal("0.00")
    monthly_principal = (
        loan.principal_amount / Decimal(loan.loan_period_months)
    ).quantize(Decimal("0.01"))
    installments_paid = (
        int(total_paid / monthly_principal) if monthly_principal > 0 else 0
    )
    next_unpaid = installments_paid + 1
    if next_unpaid > loan.loan_period_months:
        return None
    return loan.disbursement_date + relativedelta(months=next_unpaid)


# ─────────────────────────────────────────────────────────────────────────────
# Email helpers
# ─────────────────────────────────────────────────────────────────────────────


def send_loan_application_email(
    recipient_name, client_name, recipient_email, application_id, is_applicant=True
):
    dashboard_url = "https://sponsorwithpendeza.org/loans/applications/"
    subject = (
        "Your Loan Application Submitted"
        if is_applicant
        else "New Loan Application for Review"
    )

    if is_applicant:
        body = f"""
        <html><body style="font-family:Arial,sans-serif;color:#333;">
        <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #ddd;border-radius:10px;">
            <h2 style="color:#2E86C1;text-align:center;">Loan Application Submitted on Behalf of Client</h2>
            <p>Hello <strong>{recipient_name}</strong>,</p>
            <p>A loan application has been successfully submitted on behalf of <strong>{client_name}</strong>.
               Application ID: <strong>{application_id}</strong>.</p>
            <div style="text-align:center;margin:20px 0;">
                <a href="{dashboard_url}" style="background-color:#2E86C1;color:#fff;text-decoration:none;padding:10px 20px;border-radius:5px;">View Application Status</a>
            </div>
            <p style="color:#888;">- Pendeza Uganda - Finance Department</p>
        </div></body></html>
        """
    else:
        body = f"""
        <html><body style="font-family:Arial,sans-serif;color:#333;">
        <div style="max-width:600px;margin:auto;padding:20px;border:1px solid #ddd;border-radius:10px;">
            <h2 style="color:#C0392B;text-align:center;">Loan Application Approval Needed</h2>
            <p>Hello <strong>{recipient_name}</strong>,</p>
            <p>Loan application <strong>{application_id}</strong> for <strong>{client_name}</strong> awaits your review.</p>
            <div style="text-align:center;margin:20px 0;">
                <a href="{dashboard_url}" style="background-color:#C0392B;color:#fff;text-decoration:none;padding:10px 20px;border-radius:5px;">Review Application</a>
            </div>
            <p style="color:#888;">- Pendeza Uganda - Finance Department</p>
        </div></body></html>
        """

    try:
        email = EmailMultiAlternatives(
            subject, strip_tags(body), settings.EMAIL_HOST_USER, [recipient_email]
        )
        email.attach_alternative(body, "text/html")
        email.send()
        return True
    except Exception as e:
        logger.error("Error sending email to %s: %s", recipient_email, e)
        return False


def _send_html_email(subject, html_body, to):
    recipients = to if isinstance(to, list) else [to]
    send_html_email_task.delay(subject, html_body, recipients)


def send_email_to_boo(loan: Loan):
    _send_html_email(
        subject=f"Loan {loan.id} Rejected by HOF",
        html_body=f"""
        <html><body style="font-family:Arial,sans-serif;background:#f4f4f9;padding:20px;">
        <div style="max-width:600px;margin:0 auto;background:#fff;padding:30px;border-radius:8px;">
            <h2 style="color:#2c3e50;text-align:center;">Loan Rejection Notification</h2>
            <p>Dear Team,</p>
            <p>Loan <strong>{loan.id}</strong> for <strong>{loan.borrower.full_name}</strong>
               (UGX {loan.principal_amount:,.2f}) has been rejected by the Head of Finance.</p>
        </div></body></html>
        """,
        to=[settings.BOO_EMAIL],
    )


def send_email_to_boo_and_hof(loan: Loan):
    _send_html_email(
        subject=f"Loan {loan.id} Rejected by ED",
        html_body=f"""
        <html><body style="font-family:Arial,sans-serif;background:#f4f4f9;padding:20px;">
        <div style="max-width:600px;margin:0 auto;background:#fff;padding:30px;border-radius:8px;">
            <h2 style="color:#2c3e50;text-align:center;">Loan Rejection Notification</h2>
            <p>Dear Team,</p>
            <p>Loan <strong>{loan.id}</strong> for <strong>{loan.borrower.full_name}</strong>
               (UGX {loan.principal_amount:,.2f}) has been rejected by the Executive Director.</p>
        </div></body></html>
        """,
        to=[settings.BOO_EMAIL, settings.HOF_EMAIL],
    )


# ─────────────────────────────────────────────────────────────────────────────
# Loan Application views
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@admin_or_manager_or_staff_required
def loan_applications_view(request):
    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "").strip()
    sort_by = request.GET.get("sort", "-created_at").strip()
    allowed_statuses = ["pending", "boo_approved", "hof_approved"]

    qs = Loan.objects.select_related("borrower", "applied_by").filter(
        status__in=allowed_statuses
    )

    if search_query:
        qs = qs.filter(
            Q(id__icontains=search_query)
            | Q(borrower__full_name__icontains=search_query)
            | Q(borrower__reg_number__icontains=search_query)
            | Q(borrower__mobile_telephone__icontains=search_query)
            | Q(applied_by__username__icontains=search_query)
        )

    if status_filter in allowed_statuses:
        qs = qs.filter(status=status_filter)

    if request.user.profile.role in ["staff", "guest"]:
        qs = qs.filter(applied_by=request.user)

    if sort_by not in {
        "-created_at",
        "created_at",
        "-principal_amount",
        "principal_amount",
        "borrower__full_name",
    }:
        sort_by = "-created_at"
    qs = qs.order_by(sort_by, "-id")

    summary_qs = qs
    status_counts = {
        "pending": summary_qs.filter(status="pending").count(),
        "boo_approved": summary_qs.filter(status="boo_approved").count(),
        "hof_approved": summary_qs.filter(status="hof_approved").count(),
    }
    total_principal = summary_qs.aggregate(total=Sum("principal_amount"))[
        "total"
    ] or Decimal("0.00")

    loans = paginate_queryset(qs, request.GET.get("page"))
    return render(
        request,
        "loans/loan_applications.html",
        {
            "loans": loans,
            "page_obj": loans,
            "table_title": "Loan Applications",
            "search_query": search_query,
            "status_filter": status_filter,
            "status_counts": status_counts,
            "total_applications": summary_qs.count(),
            "total_principal": total_principal,
            "sort_by": sort_by,
            "status_choices": [
                ("", "All workflow stages"),
                ("pending", "Pending BOO review"),
                ("boo_approved", "Pending HOF review"),
                ("hof_approved", "Pending ED review"),
            ],
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_applications_all_view(request):
    search_query = request.GET.get("search", "")
    sort_by = request.GET.get("sort", "id")
    show_bad = request.GET.get("bad", "false").lower() == "true"

    qs = get_loan_queryset(search_query)

    if request.user.profile.role in ["staff", "guest"]:
        qs = qs.filter(applied_by=request.user)

    if show_bad:
        qs = qs.filter(
            Q(borrower_id__isnull=True)
            | ~Q(borrower_id__in=Client.objects.values_list("id", flat=True))
        )

    if sort_by in ["id", "-id", "borrower_id", "-borrower_id"]:
        qs = qs.order_by(sort_by)

    loans = paginate_queryset(qs, request.GET.get("page"), per_page=100)
    return render(
        request,
        "loans/loan_applications_all.html",
        {
            "loans": loans,
            "page_obj": loans,
            "table_title": (
                "Bad Loan Applications" if show_bad else "All Loan Applications"
            ),
            "search_query": search_query,
            "current_sort": sort_by,
            "show_bad": show_bad,
            "report_summary": loan_queryset_report_summary(qs),
        },
    )


@login_required
def loan_apply(request):
    form = LoanApplicationForm(request.POST or None, user=request.user)
    user = request.user
    role = getattr(user.profile, "role", "guest")

    if request.method == "POST" and form.is_valid():
        try:
            with transaction.atomic():
                application = form.save(commit=False)
                application.applied_by = user
                application.applied_by_role = role
                application.created_by = user
                application.save()

            borrower = application.borrower
            client_name = borrower.get_full_name()

            try:
                send_loan_application_email_task.delay(
                    recipient_name=user.username,
                    recipient_email=user.email,
                    application_id=application.id,
                    client_name=client_name,
                    is_applicant=True,
                )
                send_loan_application_email_task.delay(
                    recipient_name="Loan Officer",
                    recipient_email=settings.BOO_EMAIL,
                    application_id=application.id,
                    client_name=client_name,
                    is_applicant=False,
                )
            except Exception:
                logger.exception(
                    "Loan application %s was saved but notification queuing failed.",
                    application.id,
                )

            resp = _json_or_message(
                request, True, "Loan application submitted successfully."
            )
            if resp:
                return resp
            messages.success(
                request,
                "Loan application submitted successfully.",
                extra_tags="bg-success",
            )
            return redirect("loans:loan_applications")

        except ValidationError as e:
            logger.error("Validation error in loan_apply: %s", e)
            resp = _json_or_message(request, False, str(e), status=400)
            return resp or redirect("loans:apply_for_loan")

        except Exception as e:
            logger.exception("Unexpected error in loan_apply: %s", e)
            resp = _json_or_message(
                request,
                False,
                "An unexpected error occurred while processing the loan application.",
                status=500,
            )
            return resp or redirect("loans:apply_for_loan")

    elif request.method == "POST":
        resp = _json_or_message(
            request, False, "Please correct the errors below.", status=400
        )
        if resp:
            return resp
        if not form.is_valid():
            messages.error(
                request, "Please correct the errors below.", extra_tags="bg-danger"
            )

    open_applications = Loan.objects.filter(
        status__in=["pending", "boo_approved", "hof_approved"]
    ).count()
    active_loans = Loan.objects.filter(status__in=Loan.ACTIVE_STATUSES).count()
    clients = Client.objects.order_by("full_name", "reg_number")

    return render(
        request,
        "loans/apply_for_loan.html",
        {
            "form": form,
            "form_title": "Loan Application Form",
            "clients": clients,
            "selected_client_id": request.POST.get("client", ""),
            "open_applications": open_applications,
            "active_loans": active_loans,
        },
    )


@login_required
def client_loan_apply(request):
    current_client = _get_self_service_client(request.user)
    if current_client is None:
        messages.error(
            request,
            "We could not match your account to a client record. Please contact staff or complete client registration.",
            extra_tags="bg-danger",
        )
        return redirect("loans:client_loan_applications")

    blocking_statuses = {
        "pending",
        "boo_approved",
        "hof_approved",
        "approved",
        *Loan.ACTIVE_STATUSES,
    }
    blocking_loan = (
        Loan.objects.filter(borrower=current_client, status__in=blocking_statuses)
        .order_by("-created_at")
        .first()
    )
    open_applications = Loan.objects.filter(
        borrower=current_client,
        status__in=["pending", "boo_approved", "hof_approved", "approved"],
    ).count()
    active_loans = Loan.objects.filter(
        borrower=current_client, status__in=Loan.ACTIVE_STATUSES
    ).count()

    if blocking_loan is not None and request.method == "POST":
        messages.error(
            request,
            "You already have a pending application or running loan. Please wait until it is completed before applying again.",
            extra_tags="bg-danger",
        )
        return redirect("loans:client_loan_applications")

    if request.method == "POST":
        form = ClientSelfServiceLoanApplicationForm(request.POST)
        document_form = LoanApplicationDocumentForm(request.POST, request.FILES)
        if form.is_valid() and document_form.is_valid():
            try:
                with transaction.atomic():
                    application = form.save(
                        commit=False, borrower=current_client, user=request.user
                    )
                    application.save()
                    document_form.save(application, uploaded_by=request.user)

                try:
                    send_loan_application_email_task.delay(
                        recipient_name=request.user.get_full_name()
                        or request.user.username,
                        recipient_email=request.user.email,
                        application_id=application.id,
                        client_name=current_client.get_full_name(),
                        is_applicant=True,
                    )
                    send_loan_application_email_task.delay(
                        recipient_name="Loan Officer",
                        recipient_email=settings.BOO_EMAIL,
                        application_id=application.id,
                        client_name=current_client.get_full_name(),
                        is_applicant=False,
                    )
                except Exception:
                    logger.exception(
                        "Self-service loan application %s was saved but notification queuing failed.",
                        application.id,
                    )

                messages.success(
                    request,
                    "Loan application submitted successfully.",
                    extra_tags="bg-success",
                )
                return redirect(
                    "loans:client_loan_application_detail", loan_id=application.id
                )
            except ValidationError as exc:
                messages.error(request, str(exc), extra_tags="bg-danger")
        else:
            messages.error(
                request, "Please correct the errors below.", extra_tags="bg-danger"
            )
    else:
        form = ClientSelfServiceLoanApplicationForm()
        document_form = LoanApplicationDocumentForm()

    return render(
        request,
        "loans/client_apply_for_loan.html",
        {
            "form": form,
            "document_form": document_form,
            "form_title": "Apply for a Loan",
            "client": current_client,
            "open_applications": open_applications,
            "active_loans": active_loans,
            "blocking_loan": blocking_loan,
        },
    )


@login_required
def client_loan_applications(request):
    qs, current_client = _self_service_loan_queryset(request.user)
    summary = {
        "total_applications": 0,
        "open_applications": 0,
        "active_loans": 0,
        "total_requested": Decimal("0.00"),
        "total_documents": 0,
    }
    if current_client is None:
        messages.warning(
            request,
            "We could not match your account to a client record. Please contact staff or complete client registration.",
            extra_tags="bg-warning",
        )
        loans = []
    else:
        summary_qs = qs
        summary["total_applications"] = summary_qs.count()
        summary["open_applications"] = summary_qs.filter(
            status__in=["pending", "boo_approved", "hof_approved", "approved"]
        ).count()
        summary["active_loans"] = summary_qs.filter(
            status__in=Loan.ACTIVE_STATUSES
        ).count()
        summary["total_requested"] = summary_qs.aggregate(
            total=Sum("principal_amount")
        )["total"] or Decimal("0.00")
        loans_qs = qs.annotate(document_count=Count("documents"))
        summary["total_documents"] = (
            loans_qs.aggregate(total=Sum("document_count"))["total"] or 0
        )
        loans = paginate_queryset(loans_qs, request.GET.get("page"), per_page=20)

    return render(
        request,
        "loans/client_loan_applications.html",
        {
            "client": current_client,
            "loans": loans,
            "page_obj": loans,
            "summary": summary,
            "form_title": "My Loan Applications",
        },
    )


@login_required
def client_loan_application_detail(request, loan_id):
    qs, current_client = _self_service_loan_queryset(request.user)
    if current_client is None:
        messages.error(
            request,
            "We could not match your account to a client record. Please contact staff or complete client registration.",
            extra_tags="bg-danger",
        )
        return redirect("loans:client_loan_applications")

    loan = get_object_or_404(qs, id=loan_id)
    repayments = loan.repayments.all()
    totals = repayments.aggregate(
        total_principal=Sum("principal_payment"),
        total_interest=Sum("interest_payment"),
        total_penalty=Sum("penalty_payment"),
    )
    balances = loan.calculate_remaining_balances()
    total_outstanding_balance = sum(balances.values())
    payment_schedule = loan.generate_payment_schedule()
    schedule_total_principal = sum(
        (item["principal_payment"] for item in payment_schedule), Decimal("0.00")
    )
    schedule_total_interest = sum(
        (item["interest_payment"] for item in payment_schedule), Decimal("0.00")
    )
    schedule_total_payment = sum(
        (item["total_payment"] for item in payment_schedule), Decimal("0.00")
    )
    final_due_date = (
        payment_schedule[-1]["payment_due_date"] if payment_schedule else None
    )
    schedule_unavailable_reason = ""

    if not loan.disbursement_date:
        schedule_unavailable_reason = (
            "Your repayment schedule will be available after the loan is disbursed."
        )
    elif not loan.loan_period_months or loan.loan_period_months <= 0:
        schedule_unavailable_reason = "Your repayment schedule cannot be generated because the loan period is missing."

    return render(
        request,
        "loans/client_loan_application_detail.html",
        {
            "client": current_client,
            "loan": loan,
            "documents": loan.documents.all(),
            "repayments": repayments,
            "payment_schedule": payment_schedule,
            "schedule_total_principal": schedule_total_principal,
            "schedule_total_interest": schedule_total_interest,
            "schedule_total_payment": schedule_total_payment,
            "final_due_date": final_due_date,
            "schedule_unavailable_reason": schedule_unavailable_reason,
            "remaining_principal": balances["principal_balance"],
            "remaining_interest": balances["interest_balance"],
            "remaining_penalty": balances["penalty_balance"],
            "total_outstanding_balance": total_outstanding_balance,
            "total_principal": totals["total_principal"] or Decimal("0.00"),
            "total_interest": totals["total_interest"] or Decimal("0.00"),
            "total_penalty": totals["total_penalty"] or Decimal("0.00"),
            "form_title": f"Loan Application {loan.id}",
        },
    )


@admin_or_manager_required
def update_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    form = LoanApplicationUpdateForm(request.POST or None, instance=loan)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(
                request, "Loan details updated successfully.", extra_tags="bg-success"
            )
            return redirect("loans:loan_applications")
        messages.error(
            request, "Please correct the errors below.", extra_tags="bg-danger"
        )

    return render(
        request,
        "loans/loan_update.html",
        {
            "form": form,
            "loan": loan,
            "form_title": "Update Loan Details",
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Repayment schedule
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@admin_or_manager_or_staff_required
def repayment_schedule(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    schedule = loan.generate_payment_schedule()
    schedule_unavailable_reason = ""

    if not loan.disbursement_date:
        schedule_unavailable_reason = (
            "Repayment schedule will be available after this loan is disbursed."
        )
    elif not loan.loan_period_months or loan.loan_period_months <= 0:
        schedule_unavailable_reason = (
            "Repayment schedule cannot be generated because the loan period is missing."
        )

    # First payment gives monthly figures (same for all — flat rate)
    first = schedule[0] if schedule else {}
    monthly_principal = first.get("principal_payment", 0)
    monthly_interest = first.get("interest_payment", 0)
    monthly_payment = first.get("total_payment", 0)

    return render(
        request,
        "loans/repayment_schedule.html",
        {
            "loan": loan,
            "repayment_schedule": schedule,
            "monthly_principal_repayment": monthly_principal,
            "monthly_interest_repayment": monthly_interest,
            "monthly_payment": monthly_payment,
            "total_interest": loan.total_interest,
            "total_cost_of_loan": loan.total_repayable,  # uses model @property
            "loan_period_years": (
                loan.loan_period_months / 12 if loan.loan_period_months else 0
            ),
            "interest_method": loan.interest_method,
            "schedule_unavailable_reason": schedule_unavailable_reason,
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Loan list views
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@admin_or_manager_or_staff_required
def disbursed_loans_view(request):
    qs = get_loan_queryset(request.GET.get("search")).filter(
        status__in=["disbursed", "overdue", "repaid"]
    )
    # Evaluate once for totals, then paginate
    total_disbursed = Decimal("0.00")
    total_interest_all = Decimal("0.00")
    for loan in qs:
        total_disbursed += loan.principal_amount or Decimal("0")
        total_interest_all += loan.total_interest or Decimal("0")

    loans = paginate_queryset(qs, request.GET.get("page"))

    loans_with_info = [
        {
            "loan_id": loan.id,
            "borrower": loan.borrower,
            "principal_amount": loan.principal_amount,
            "total_interest": loan.total_interest,
            "interest_rate": loan.interest_rate,
            "loan_period_months": loan.loan_period_months,
            "start_date": loan.start_date,
            "due_date": loan.due_date,
            "status": loan.get_status_display(),
            "disbursement_date": loan.disbursement_date,
            "account_number": loan.account.account_number if loan.account else None,
            "payment_method": disbursement.payment_method,
        }
        for loan in loans
        for disbursement in loan.disbursements.all()
    ]

    return render(
        request,
        "loans/disbursed_loans_list.html",
        {
            "loans_with_disbursement_info": loans_with_info,
            "loans": loans,
            "table_title": "Disbursed Loans",
            "total_disbursed": total_disbursed,
            "total_interest_all": total_interest_all,
            "total_recoverable": total_disbursed + total_interest_all,
            "loan_count": qs.count(),
            "search_query": request.GET.get("search", ""),
        },
    )


@login_required
@admin_or_manager_or_staff_required
def approved_loans_view(request):
    qs = get_loan_queryset(request.GET.get("search")).filter(status="approved")
    if request.user.profile.role in ["staff", "guest"]:
        qs = qs.filter(applied_by=request.user)
    return render(
        request,
        "loans/approved_loans_list.html",
        {
            "loans": paginate_queryset(qs, request.GET.get("page")),
            "table_title": "Pending Loan Disbursements",
            "search_query": request.GET.get("search"),
            "report_summary": loan_queryset_report_summary(qs),
        },
    )


@login_required
@admin_or_manager_or_staff_required
def rejected_loans_view(request):
    qs = get_loan_queryset(request.GET.get("search")).filter(
        status__in=["ed_rejected", "hof_rejected"]
    )
    if request.user.profile.role in ["staff", "guest"]:
        qs = qs.filter(applied_by=request.user)
    return render(
        request,
        "loans/rejected_loans_list.html",
        {
            "loans": paginate_queryset(qs, request.GET.get("page")),
            "table_title": "Rejected Loans",
            "search_query": request.GET.get("search"),
            "report_summary": loan_queryset_report_summary(qs),
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Disbursement
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@admin_or_manager_required
def disburse_loan(request):
    form = LoanDisbursementForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        disbursement = form.save(commit=False)
        loan = form.cleaned_data["loan"]
        disbursement.loan = loan
        disbursement_date = form.cleaned_data.get("disbursement_date")
        try:
            with transaction.atomic():
                loan.disburse(disbursement_date)
                disbursement.save()  # triggers journal entries on insert
        except ValidationError as exc:
            messages.error(request, str(exc), extra_tags="bg-danger")
            return redirect("loans:disburse_loan")
        except Exception:
            logger.exception("Loan %s disbursement failed.", loan.id)
            messages.error(
                request,
                "The disbursement could not be posted. No loan or journal changes were saved.",
                extra_tags="bg-danger",
            )
            return redirect("loans:disburse_loan")

        messages.success(
            request,
            f"Loan ID {loan.id} disbursed successfully.",
            extra_tags="bg-success",
        )
        return redirect("loans:disburse_loan")

    if request.method == "POST":
        messages.error(
            request, "Please check the form for errors.", extra_tags="bg-danger"
        )

    return render(
        request,
        "loans/disburse_loan.html",
        {
            "approved_loans": Loan.objects.filter(status="approved"),
            "form_title": "Disburse Approved Loans",
            "form": form,
        },
    )


@login_required
@admin_or_manager_required
def disburse_all_loans(request):
    eligible, ineligible = [], []

    for loan in Loan.objects.filter(status="approved"):
        has_balance = any(
            sum(rl.calculate_remaining_balances().values()) > 0
            for rl in Loan.objects.filter(
                borrower=loan.borrower, status__in=["disbursed", "overdue"]
            ).exclude(id=loan.id)
        )
        (ineligible if has_balance else eligible).append(loan)

    if not eligible:
        messages.warning(
            request,
            "No approved loans available for disbursement.",
            extra_tags="bg-warning",
        )
        return redirect("loans:disbursed_loans")

    form = LoanAllDisbursementForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        try:
            count = form.save(eligible)
        except ValidationError as exc:
            messages.error(request, str(exc), extra_tags="bg-danger")
            return redirect("loans:disburse_all_loans")
        except Exception:
            logger.exception("Bulk loan disbursement failed.")
            messages.error(
                request,
                "Bulk disbursement failed. No partial disbursements were saved.",
                extra_tags="bg-danger",
            )
            return redirect("loans:disburse_all_loans")
        messages.success(
            request, f"{count} loans disbursed successfully.", extra_tags="bg-success"
        )
        if ineligible:
            messages.warning(
                request,
                f"{len(ineligible)} loans skipped due to existing running balances.",
                extra_tags="bg-warning",
            )
        return redirect("loans:disburse_all_loans")

    if request.method == "POST":
        messages.error(
            request, "Please check the form for errors.", extra_tags="bg-danger"
        )

    return render(
        request,
        "loans/disburse_all_loans.html",
        {
            "form_title": "Disburse All Approved Loans",
            "form": form,
            "eligible_loans": eligible,
            "ineligible_loans": ineligible,
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Approval chain
# ─────────────────────────────────────────────────────────────────────────────

# Approval transition table: current_status → (new_status, approved_by_field, notify_to)
_APPROVAL_TRANSITIONS = {
    ("pending", "boo"): ("boo_approved", "approved_by_boo"),
    ("boo_approved", "hof"): ("hof_approved", "approved_by_hof"),
    ("hof_approved", "ed"): ("approved", "approved_by_ed"),
}


@login_required
def approve_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    user = request.user

    try:
        new_status = loan.approve(user)
    except ValidationError:
        messages.error(
            request,
            "You are not authorized to approve this loan at this stage.",
            extra_tags="bg-danger",
        )
        return redirect("loans:loan_applications")

    base_url = request.build_absolute_uri("/")
    approver_name = user.get_full_name() or user.username

    def queue_approval_notification():
        try:
            send_loan_approval_notification_task.delay(
                loan.id,
                new_status,
                approver_name,
                base_url,
            )
        except Exception:
            logger.exception(
                "Loan %s was approved as %s, but approval notification queuing failed.",
                loan.id,
                new_status,
            )

    transaction.on_commit(queue_approval_notification)

    messages.success(
        request, f"Loan {loan.id} approved ({new_status}).", extra_tags="bg-success"
    )
    return redirect("loans:loan_applications")


@login_required
@admin_or_manager_or_staff_required
def approve_all_loans(request):
    user = request.user
    role = user.profile.role

    stage_map = {
        "boo": "pending",
        "hof": "boo_approved",
        "ed": "hof_approved",
    }

    if role not in stage_map:
        messages.error(request, "You are not authorized.", extra_tags="bg-danger")
        return redirect("loans:loan_applications")

    pending_status = stage_map[role]
    pending_loans = Loan.objects.filter(status=pending_status)

    if not pending_loans.exists():
        messages.info(
            request,
            f"No {pending_status.replace('_', ' ').title()} loans to approve.",
            extra_tags="bg-info",
        )
        return redirect("loans:loan_applications")

    for loan in pending_loans:
        loan.approve(user)

    messages.success(
        request, f"All {pending_status} loans approved.", extra_tags="bg-success"
    )
    return redirect("loans:loan_applications")


@login_required
@admin_or_manager_required
def reject_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)

    if loan.status == "approved":
        messages.error(
            request,
            f"Loan {loan.id} cannot be rejected — it is already approved.",
            extra_tags="bg-warning",
        )
        return redirect("loans:loan_applications")

    rejection_map = {
        "pending": "rejected",
        "boo_approved": "hof_rejected",
        "hof_approved": "ed_rejected",
    }

    if loan.status not in rejection_map:
        messages.error(
            request,
            f"Loan {loan.id} cannot be rejected at status '{loan.status}'.",
            extra_tags="bg-warning",
        )
        return redirect("loans:loan_applications")

    try:
        loan.reject()
    except ValidationError:
        messages.error(
            request,
            f"Loan {loan.id} cannot be rejected at status '{loan.status}'.",
            extra_tags="bg-warning",
        )
        return redirect("loans:loan_applications")

    if loan.status == "hof_rejected":
        send_email_to_boo(loan)
    else:
        send_email_to_boo_and_hof(loan)

    messages.info(request, f"Loan {loan.id} rejected.", extra_tags="bg-danger")
    return redirect("loans:loan_applications")


@login_required
@admin_or_manager_required
def delete_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id)
    try:
        loan_ref = str(loan)
        loan.delete()
        messages.success(
            request, f"{loan_ref} deleted successfully!", extra_tags="bg-danger"
        )
    except Exception as e:
        logger.error("Error deleting loan %s: %s", loan_id, e)
        messages.error(
            request, "An error occurred during deletion.", extra_tags="bg-danger"
        )
    return redirect("loans:loan_applications")


# ─────────────────────────────────────────────────────────────────────────────
# Repayment & penalty
# ─────────────────────────────────────────────────────────────────────────────


def _loans_with_balance(status_list):
    """
    Returns loans in the given statuses that still have an outstanding balance,
    with remaining_* attributes attached. Used by repayment and penalty views.
    """
    result = []
    for loan in (
        Loan.objects.filter(status__in=status_list)
        .select_related("borrower")
        .prefetch_related("repayments", "penalties")
    ):
        if not getattr(loan, "borrower", None):
            continue
        b = loan.calculate_remaining_balances()
        if any(v > 0 for v in b.values()):
            loan.remaining_principal = b["principal_balance"]
            loan.remaining_interest = b["interest_balance"]
            loan.remaining_penalty = b["penalty_balance"]
            result.append(loan)
    return result


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def loan_repayment_create_view(request):
    loans = _loans_with_balance(["disbursed", "overdue"])

    if request.method == "POST":
        form = LoanRepaymentForm(request.POST)
        if form.is_valid():
            repayment = form.save(commit=False)
            repayment.loan = form.cleaned_data["loan"]
            if not getattr(repayment.loan, "borrower", None):
                messages.error(request, "This loan has no valid borrower attached.")
                return redirect("loans:loan_repayment_create")
            try:
                repayment.save()  # clean() + _post_entries() + update_status() in model
            except ValidationError as exc:
                form.add_error(None, exc)
                messages.error(
                    request, "Please correct the errors below.", extra_tags="bg-danger"
                )
            except Exception:
                logger.exception(
                    "Repayment posting failed for loan %s.", repayment.loan_id
                )
                messages.error(
                    request,
                    "The repayment could not be posted. No repayment or journal changes were saved.",
                    extra_tags="bg-danger",
                )
            else:
                messages.success(
                    request,
                    "Loan repayment submitted successfully.",
                    extra_tags="bg-success",
                )
                return redirect("loans:loan_repayment_create")
        messages.error(
            request, "Please correct the errors below.", extra_tags="bg-danger"
        )
    else:
        form = LoanRepaymentForm()

    return render(
        request,
        "loans/loan_repayment_form.html",
        {
            "form": form,
            "form_title": "Repay Loans",
            "loans": loans,
        },
    )


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def loan_penalty_create_view(request):
    loans = _loans_with_balance(["disbursed", "overdue"])

    if request.method == "POST":
        form = LoanPenaltyForm(request.POST, user=request.user)
        if form.is_valid():
            penalty = form.save(commit=False)
            penalty.created_by = request.user
            try:
                penalty.save()  # _post_entries() + update_status() in model
            except ValidationError as exc:
                form.add_error(None, exc)
                messages.error(
                    request, "Please correct the errors below.", extra_tags="bg-danger"
                )
            except Exception:
                logger.exception("Penalty posting failed for loan %s.", penalty.loan_id)
                messages.error(
                    request,
                    "The penalty could not be posted. No penalty or journal changes were saved.",
                    extra_tags="bg-danger",
                )
            else:
                messages.success(
                    request,
                    f"Penalty of {penalty.penalty_amount:,.2f} added to Loan {penalty.loan.id}.",
                    extra_tags="bg-success",
                )
                return redirect("loans:loan_penalty_create")
        if not form.is_valid():
            messages.error(
                request, "Please correct the errors below.", extra_tags="bg-danger"
            )
    else:
        form = LoanPenaltyForm(user=request.user)

    return render(
        request,
        "loans/loan_penalty_form.html",
        {
            "form": form,
            "form_title": "Add Loan Penalty",
            "loans": loans,
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_detail_view(request, loan_id):
    loan = get_object_or_404(
        Loan.objects.select_related("borrower", "account").prefetch_related(
            "repayments", "documents"
        ),
        id=loan_id,
    )
    balances = loan.calculate_remaining_balances()
    repayments = loan.repayments.all()
    totals = repayments.aggregate(
        total_principal=Sum("principal_payment"),
        total_interest=Sum("interest_payment"),
        total_penalty=Sum("penalty_payment"),
    )
    total_remaining_balance = sum(balances.values())
    documents_page = Paginator(loan.documents.all(), 5).get_page(
        request.GET.get("documents_page")
    )
    repayments_page = Paginator(repayments, 8).get_page(
        request.GET.get("repayments_page")
    )

    return render(
        request,
        "loans/loan_detail.html",
        {
            "loan": loan,
            "remaining_principal": balances["principal_balance"],
            "remaining_interest": balances["interest_balance"],
            "remaining_penalty": balances["penalty_balance"],
            "total_remaining_balance": total_remaining_balance,
            "documents": documents_page,
            "documents_page": documents_page,
            "document_upload_form": StaffLoanApplicationDocumentForm(),
            "repayments": repayments_page,
            "repayments_page": repayments_page,
            "borrower_name": loan.borrower.full_name,
            "total_principal": totals["total_principal"] or 0,
            "total_interest": totals["total_interest"] or 0,
            "total_penalty": totals["total_penalty"] or 0,
            "form_title": (
                f"{loan.borrower.full_name} | Loan id: ({loan.id}) "
                f"| Reg No: {loan.borrower.reg_number}"
            ),
        },
    )


@login_required
def client_loan_application_document_open(request, loan_id, document_id):
    qs, current_client = _self_service_loan_queryset(request.user)
    if current_client is None:
        messages.error(
            request,
            "We could not match your account to a client record. Please contact staff or complete client registration.",
            extra_tags="bg-danger",
        )
        return redirect("loans:client_loan_applications")

    loan = get_object_or_404(qs, id=loan_id)
    document = get_object_or_404(
        LoanApplicationDocument.objects.filter(loan=loan),
        id=document_id,
    )
    if not document.file:
        raise Http404("Document file not found.")

    return redirect(document.file.url)


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def upload_loan_application_document(request, loan_id):
    loan = get_object_or_404(Loan.objects.select_related("borrower"), id=loan_id)
    if request.method != "POST":
        return redirect("loans:loan_detail", loan_id=loan.id)

    form = StaffLoanApplicationDocumentForm(request.POST, request.FILES)
    if form.is_valid():
        form.save(loan=loan, uploaded_by=request.user)
        messages.success(
            request,
            "Loan application document uploaded successfully.",
            extra_tags="bg-success",
        )
    else:
        for field_errors in form.errors.values():
            for error in field_errors:
                messages.error(request, error, extra_tags="bg-danger")

    return redirect("loans:loan_detail", loan_id=loan.id)


@login_required
@admin_or_manager_required
def delete_repayment(request, repayment_id):
    repayment = get_object_or_404(LoanRepayment, id=repayment_id)
    if request.method == "POST":
        linked_transactions = TransactionHistory.objects.filter(
            loan=repayment.loan,
            transaction_date=repayment.repayment_date,
        ).filter(
            Q(description__icontains=f"Loan {repayment.loan_id}")
            | Q(description__icontains="Loan repayment")
            | Q(description__icontains="Interest received")
            | Q(description__icontains="Penalty payment")
        )
        if linked_transactions.exists():
            messages.error(
                request,
                "This repayment has posted journal entries and cannot be deleted. Post a correcting entry instead.",
                extra_tags="bg-danger",
            )
        else:
            repayment.delete()
            repayment.loan.update_status()
            messages.success(
                request, "Repayment deleted successfully.", extra_tags="bg-success"
            )
    return redirect(request.META.get("HTTP_REFERER", "loans:loan_list"))


# ─────────────────────────────────────────────────────────────────────────────
# Chart of Accounts
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@admin_or_manager_or_staff_required
def chart_of_accounts_list_view(request):
    accounts_by_type = {}
    for account in ChartOfAccounts.objects.all():
        accounts_by_type.setdefault(account.get_account_type_display(), []).append(
            account
        )
    return render(
        request,
        "loans/chart_of_accounts_list.html",
        {
            "accounts_by_type": accounts_by_type,
            "table_title": "Chart of Accounts",
        },
    )


@login_required
@admin_or_manager_required
def add_chart_of_account_view(request):
    form = ChartOfAccountsForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(
            request, "Account added successfully!", extra_tags="bg-success"
        )
        return redirect("loans:add_chart_of_account")
    return render(
        request,
        "loans/chart_of_account_add.html",
        {
            "form": form,
            "table_title": "Add New Account",
        },
    )


@login_required
@admin_or_manager_or_staff_required
@transaction.atomic
def chart_of_account_update_view(request, account_id):
    account = get_object_or_404(ChartOfAccounts, id=account_id)
    form = ChartOfAccountsForm(request.POST or None, instance=account)

    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f"Account '{account.account_name}' updated.",
                extra_tags="bg-success",
            )
            return redirect("loans:chart_of_accounts_list")
        messages.error(request, "Error updating account.", extra_tags="bg-danger")

    return render(
        request,
        "loans/chart_of_account_update.html",
        {
            "form": form,
            "account": account,
            "page_title": "Update Account",
        },
    )


@login_required
@admin_required
@transaction.atomic
def chart_of_account_delete_view(request, account_id):
    account = get_object_or_404(ChartOfAccounts, id=account_id)
    try:
        name = account.account_name
        account.delete()
        messages.success(request, f"Account '{name}' deleted.", extra_tags="bg-success")
    except Exception as e:
        logger.error("Error deleting account %s: %s", account_id, e)
        messages.error(request, "Error during deletion.", extra_tags="bg-danger")
    return redirect("loans:chart_of_accounts_list")


# ─────────────────────────────────────────────────────────────────────────────
# Excel imports
# ─────────────────────────────────────────────────────────────────────────────


@login_required
@admin_required
@transaction.atomic
def import_coa_data(request):
    form = ImportCOAForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        f = request.FILES.get("excel_file")
        if f and f.name.endswith(".xlsx"):
            errors = _import_accounts(f)
            for e in errors:
                messages.error(request, e, extra_tags="bg-danger")
            if not errors:
                messages.success(
                    request, "Accounts imported successfully!", extra_tags="bg-success"
                )
            return redirect("loans:chart_of_accounts_list")
        messages.error(
            request, "Please upload a valid .xlsx file.", extra_tags="bg-danger"
        )

    return render(
        request,
        "loans/accounts_import.html",
        {
            "form_name": "Import Accounts - Excel",
            "form": form,
        },
    )


@transaction.atomic
def _import_accounts(excel_file):
    errors = []
    try:
        sheet = load_workbook(excel_file).active
        valid_types = set(dict(ChartOfAccounts.ACCOUNT_TYPE_CHOICES))

        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            name, acct_type, acct_number, desc = (
                row[0].value,
                row[1].value,
                row[2].value,
                row[3].value,
            )
            if acct_number is None:
                errors.append(f"Row {row_num}: missing account number")
                continue
            acct_number = str(acct_number)
            if not (name and acct_type and acct_number):
                errors.append(f"Row {row_num}: missing required fields")
                continue
            if acct_type not in valid_types:
                errors.append(f"Row {row_num}: invalid account type '{acct_type}'")
                continue
            if not acct_number.isdigit():
                errors.append(f"Row {row_num}: account number must be numeric")
                continue
            try:
                ChartOfAccounts.objects.create(
                    account_name=name,
                    account_type=acct_type,
                    account_number=acct_number,
                    description=desc,
                )
            except Exception as e:
                errors.append(f"Row {row_num}: {e}")
    except Exception as e:
        errors.append(f"Failed to process file: {e}")
    return errors


@login_required
@admin_required
def import_loan_data(request):
    form = ImportLoansForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        f = request.FILES.get("excel_file")
        if f and f.name.endswith(".xlsx"):
            errors = _import_loans(f)
            for e in errors:
                messages.error(request, e, extra_tags="bg-danger")
            if not errors:
                messages.success(
                    request, "Loans imported successfully!", extra_tags="bg-success"
                )
            return redirect("loans:loan_applications")
        messages.error(
            request, "Please upload a valid .xlsx file.", extra_tags="bg-danger"
        )

    return render(
        request,
        "loans/import_loans.html",
        {
            "form_name": "Import Loans - Excel",
            "form": form,
        },
    )


def _import_loans(excel_file):
    errors = []
    try:
        sheet = load_workbook(excel_file).active
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            (
                reg_number,
                full_name,
                picture,
                mobile_telephone,
                principal_amount,
                interest_rate,
                start_date,
                loan_period_months,
                interest_method,
            ) = (c.value for c in row[:9])
            if not full_name:
                errors.append(f"Row {row_num}: missing full name")
                continue
            try:
                client, _ = Client.objects.get_or_create(
                    reg_number=reg_number,
                    defaults={
                        "full_name": full_name,
                        "picture": picture,
                        "mobile_telephone": mobile_telephone,
                    },
                )
                Loan.objects.create(
                    borrower=client,
                    principal_amount=principal_amount,
                    interest_rate=interest_rate,
                    start_date=start_date,
                    loan_period_months=loan_period_months,
                    interest_method=interest_method,
                )
            except Exception as e:
                errors.append(f"Row {row_num}: {e}")
    except Exception as e:
        errors.append(f"Failed to process file: {e}")
    return errors


# ─────────────────────────────────────────────────────────────────────────────
# Reports
# ─────────────────────────────────────────────────────────────────────────────


def get_financial_year_dates():
    today = date.today()
    if today.month >= 7:
        return date(today.year, 7, 1), date(today.year + 1, 6, 30)
    return date(today.year - 1, 7, 1), date(today.year, 6, 30)


@login_required
@admin_or_manager_or_staff_required
def ledger_report_view(request):
    accounts = ChartOfAccounts.objects.all()
    selected_account_id = request.GET.get("account_id")
    fy_start, fy_end = get_financial_year_dates()
    start_date = request.GET.get("start_date") or fy_start
    end_date = request.GET.get("end_date") or fy_end

    ledger_data = []
    total_debits = total_credits = opening_balance = 0
    selected_account = None

    if selected_account_id:
        selected_account = get_object_or_404(ChartOfAccounts, id=selected_account_id)

        # Opening balance: all entries before the range
        for txn in TransactionHistory.objects.filter(
            account=selected_account, transaction_date__lt=start_date
        ):
            opening_balance += (
                txn.amount if txn.transaction_type == "debit" else -txn.amount
            )

        running = opening_balance
        for txn in TransactionHistory.objects.filter(
            account=selected_account,
            transaction_date__range=[start_date, end_date],
        ).order_by("transaction_date"):
            if txn.transaction_type == "debit":
                txn.debit = txn.amount
                txn.credit = 0
                total_debits += txn.amount
            else:
                txn.debit = 0
                txn.credit = txn.amount
                total_credits += txn.amount
            running += txn.debit - txn.credit
            txn.running_balance = running
            ledger_data.append(txn)

    return render(
        request,
        "loans/ledger_report.html",
        {
            "ledger_data": ledger_data,
            "accounts": accounts,
            "selected_account": selected_account,
            "selected_account_id": selected_account_id,
            "start_date": start_date,
            "end_date": end_date,
            "total_debits": total_debits,
            "total_credits": total_credits,
            "opening_balance": opening_balance,
            "closing_balance": opening_balance + total_debits - total_credits,
            "transaction_count": len(ledger_data),
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_aging_report(request):
    today = timezone.now().date()

    bucket_keys = [
        "Current",
        "1-30 days overdue",
        "31-60 days overdue",
        "61-90 days overdue",
        "91-180 days overdue",
        "Over 180 days overdue",
    ]
    aging_buckets = {k: [] for k in bucket_keys}
    zero_totals = lambda: {
        "total_principal": Decimal("0.00"),
        "total_principal_due": Decimal("0.00"),
        "total_interest_due": Decimal("0.00"),
        "total_penalty_due": Decimal("0.00"),
        "total_outstanding_balance": Decimal("0.00"),
        "total_paid": Decimal("0.00"),
    }
    bucket_totals = {k: zero_totals() for k in bucket_keys}
    grand_totals = zero_totals()

    loans = (
        Loan.objects.filter(
            status__in=["disbursed", "overdue"],
            disbursement_date__isnull=False,
        )
        .select_related("borrower")
        .prefetch_related("repayments", "penalties")
    )

    for loan in loans:
        try:
            balances = loan.calculate_remaining_balances()
            outstanding = sum(balances.values())
            if outstanding <= 0:
                continue

            aging = compute_installment_based_days_overdue(loan, today)

            paid_all = loan.repayments.aggregate(
                total_principal=Sum("principal_payment"),
                total_interest=Sum("interest_payment"),
                total_penalty=Sum("penalty_payment"),
            )
            last_repayment = loan.repayments.order_by("-repayment_date").first()
            total_paid_all = (
                (paid_all["total_principal"] or Decimal("0.00"))
                + (paid_all["total_interest"] or Decimal("0.00"))
                + (paid_all["total_penalty"] or Decimal("0.00"))
            )

            loan_info = {
                "loan_id": loan.id,
                "borrower": loan.borrower.full_name,
                "principal_amount": loan.principal_amount,
                "interest_rate": loan.interest_rate,
                "loan_period_months": loan.loan_period_months,
                "disbursement_date": loan.disbursement_date,
                "due_date": aging["final_due_date"],
                "next_due_date": aging["next_due_date"],
                "first_unpaid_due_date": aging["first_unpaid_due_date"],
                "days_overdue": aging["days_overdue"],
                "monthly_installment": aging["monthly_installment"],
                "total_repayable": aging["total_repayable"],
                "installments_due_by_now": aging["installments_due_by_now"],
                "total_due_by_today": aging["total_due_by_today"],
                "total_paid_pi": aging["total_paid_pi"],
                "shortfall": aging["shortfall"],
                "overdue_amount": aging["shortfall"],
                "principal_due": balances["principal_balance"],
                "interest_due": balances["interest_balance"],
                "penalty_due": balances["penalty_balance"],
                "outstanding_balance": outstanding,
                "total_paid": total_paid_all,
                "last_repayment_date": (
                    last_repayment.repayment_date if last_repayment else None
                ),
                "loan_product": loan.get_loan_purpose_display(),
                "status": loan.status,
            }

            d = aging["days_overdue"]
            if d == 0:
                bucket_key = "Current"
            elif d <= 30:
                bucket_key = "1-30 days overdue"
            elif d <= 60:
                bucket_key = "31-60 days overdue"
            elif d <= 90:
                bucket_key = "61-90 days overdue"
            elif d <= 180:
                bucket_key = "91-180 days overdue"
            else:
                bucket_key = "Over 180 days overdue"
            loan_info["aging_bucket"] = bucket_key

            aging_buckets[bucket_key].append(loan_info)
            bt = bucket_totals[bucket_key]
            bt["total_principal"] += loan.principal_amount or Decimal("0")
            bt["total_principal_due"] += balances["principal_balance"]
            bt["total_interest_due"] += balances["interest_balance"]
            bt["total_penalty_due"] += balances["penalty_balance"]
            bt["total_outstanding_balance"] += outstanding
            bt["total_paid"] += total_paid_all

        except Exception as e:
            logger.error(
                "Error processing loan %s for aging: %s", loan.id, e, exc_info=True
            )

    for bucket_key in aging_buckets:
        aging_buckets[bucket_key].sort(
            key=lambda x: (
                x["days_overdue"],
                x["next_due_date"] or today,
                x["borrower"],
            )
        )

    for bt in bucket_totals.values():
        for k in grand_totals:
            grand_totals[k] += bt[k]

    fmt = lambda v: f"{float(v):,.0f}"

    return render(
        request,
        "loans/loan_aging_report.html",
        {
            "bucket_data": [
                {
                    "key": k,
                    "loans": aging_buckets[k],
                    "totals": {
                        fk: fmt(bucket_totals[k][fk]) for fk in bucket_totals[k]
                    },
                }
                for k in bucket_keys
            ],
            "formatted_grand_totals": {k: fmt(v) for k, v in grand_totals.items()},
            "bucket_totals": bucket_totals,
            "grand_totals": grand_totals,
            "table_title": "Loan Aging Report",
            "total_loans": sum(len(v) for v in aging_buckets.values()),
            "now": timezone.now(),
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_arrears_report(request):
    today = timezone.now().date()

    bucket_keys = [
        "1-30 Days (WATCH)",
        "31-60 Days (SUBSTANDARD)",
        "61-90 Days (SUBSTANDARD)",
        "91-120 Days (DOUBTFUL)",
        "121-180 Days (DOUBTFUL)",
        "181-365 Days (LOSS)",
        "Over 365 Days (LOSS)",
    ]
    arrears_buckets = {k: [] for k in bucket_keys}
    zero_bt = lambda: {
        "loan_count": 0,
        "total_principal": Decimal("0.00"),
        "total_principal_due": Decimal("0.00"),
        "total_interest_due": Decimal("0.00"),
        "total_outstanding": Decimal("0.00"),
    }
    bucket_totals = {k: zero_bt() for k in bucket_keys}
    grand_totals = zero_bt()

    loans = (
        Loan.objects.filter(
            status__in=["disbursed", "overdue"],
            disbursement_date__isnull=False,
        )
        .select_related("borrower")
        .prefetch_related("repayments")
    )

    for loan in loans:
        try:
            balances = loan.calculate_remaining_balances()
            outstanding = sum(balances.values())
            if outstanding <= 0:
                continue

            # Reuse the shared aging helper — consistent with aging report
            aging = compute_installment_based_days_overdue(loan, today)
            days_overdue = aging["days_overdue"]

            if days_overdue == 0:
                continue  # performing — not an arrears loan

            if days_overdue <= 30:
                bucket = "1-30 Days (WATCH)"
            elif days_overdue <= 60:
                bucket = "31-60 Days (SUBSTANDARD)"
            elif days_overdue <= 90:
                bucket = "61-90 Days (SUBSTANDARD)"
            elif days_overdue <= 120:
                bucket = "91-120 Days (DOUBTFUL)"
            elif days_overdue <= 180:
                bucket = "121-180 Days (DOUBTFUL)"
            elif days_overdue <= 365:
                bucket = "181-365 Days (LOSS)"
            else:
                bucket = "Over 365 Days (LOSS)"

            arrears_buckets[bucket].append(
                {
                    "loan_id": loan.id,
                    "borrower": loan.borrower.full_name,
                    "principal_amount": loan.principal_amount,
                    "interest_rate": loan.interest_rate,
                    "term_months": loan.loan_period_months,
                    "disbursement_date": loan.disbursement_date,
                    "final_due_date": aging["final_due_date"],
                    "next_due_date": aging["next_due_date"],
                    "days_overdue": days_overdue,
                    "principal_due": balances["principal_balance"],
                    "interest_due": balances["interest_balance"],
                    "outstanding": outstanding,
                }
            )

            bt = bucket_totals[bucket]
            bt["loan_count"] += 1
            bt["total_principal"] += loan.principal_amount
            bt["total_principal_due"] += balances["principal_balance"]
            bt["total_interest_due"] += balances["interest_balance"]
            bt["total_outstanding"] += outstanding

        except Exception as e:
            logger.error("Error in arrears report for loan %s: %s", loan.id, e)

    for bucket in arrears_buckets:
        arrears_buckets[bucket].sort(
            key=lambda x: (-x["days_overdue"], -float(x["outstanding"]))
        )

    for bt in bucket_totals.values():
        for k in grand_totals:
            grand_totals[k] += bt[k]

    fmt = lambda x: f"{float(x):,.0f}"

    return render(
        request,
        "loans/loan_arrears_report.html",
        {
            "bucket_data": [
                {
                    "name": k,
                    "loans": arrears_buckets[k],
                    "count": bucket_totals[k]["loan_count"],
                    "totals": {
                        "principal": fmt(bucket_totals[k]["total_principal"]),
                        "principal_due": fmt(bucket_totals[k]["total_principal_due"]),
                        "interest_due": fmt(bucket_totals[k]["total_interest_due"]),
                        "outstanding": fmt(bucket_totals[k]["total_outstanding"]),
                    },
                }
                for k in bucket_keys
            ],
            "grand_totals": {
                "count": grand_totals["loan_count"],
                "principal": fmt(grand_totals["total_principal"]),
                "principal_due": fmt(grand_totals["total_principal_due"]),
                "interest_due": fmt(grand_totals["total_interest_due"]),
                "outstanding": fmt(grand_totals["total_outstanding"]),
            },
            "report_date": today.strftime("%d %B %Y"),
            "title": "Loan Arrears Report (Past Due Only)",
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_portfolio_report(request):
    today = timezone.now().date()
    loan_data = []
    totals = {
        k: Decimal("0.00")
        for k in [
            "total_principal",
            "total_remaining_principal",
            "total_remaining_interest",
            "total_penalty_balance",
            "total_remaining_balance",
        ]
    }

    for loan in (
        Loan.objects.select_related("borrower").prefetch_related("repayments").all()
    ):
        try:
            balances = loan.calculate_remaining_balances()
            total_bal = sum(balances.values())
            if total_bal <= 0:
                continue

            days_overdue = (
                max((today - loan.due_date).days, 0)
                if loan.due_date and loan.due_date < today
                else 0
            )
            last_repayment = loan.repayments.order_by("-repayment_date").first()
            next_payment = _get_next_payment_date(loan)

            totals["total_principal"] += loan.principal_amount
            totals["total_remaining_principal"] += balances["principal_balance"]
            totals["total_remaining_interest"] += balances["interest_balance"]
            totals["total_penalty_balance"] += balances["penalty_balance"]
            totals["total_remaining_balance"] += total_bal

            loan_data.append(
                {
                    "loan_id": loan.id,
                    "borrower": loan.borrower.full_name,
                    "principal_amount": loan.principal_amount,
                    "interest_rate": loan.interest_rate,
                    "loan_period_months": loan.loan_period_months,
                    "remaining_principal": balances["principal_balance"],
                    "remaining_interest": balances["interest_balance"],
                    "penalty_balance": balances["penalty_balance"],
                    "total_remaining_balance": total_bal,
                    "disbursement_date": loan.disbursement_date,
                    "due_date": loan.due_date,
                    "days_overdue": days_overdue,
                    "last_payment": (
                        last_repayment.repayment_date if last_repayment else None
                    ),
                    "next_payment": next_payment,
                }
            )
        except Exception as e:
            logger.error(
                "Portfolio report error for loan %s: %s", loan.id, e, exc_info=True
            )

    loan_data.sort(key=lambda x: x["disbursement_date"] or date.min)
    page_obj = paginate_queryset(loan_data, request.GET.get("page"))

    return render(
        request,
        "loans/loan_portfolio_report.html",
        {
            "page_obj": page_obj,
            "table_title": "Loan Portfolio Report",
            **totals,
        },
    )


@login_required
def portfolio_at_risk(request):
    today = timezone.now().date()
    loans = (
        Loan.objects.filter(
            status__in=["disbursed", "overdue"], disbursement_date__isnull=False
        )
        .select_related("borrower")
        .prefetch_related("repayments")
    )

    par = {
        k: Decimal("0.00")
        for k in ["par_1", "par_30", "par_60", "par_90", "par_120", "par_180"]
    }
    total_portfolio = Decimal("0.00")
    total_loans_count = 0

    for loan in loans:
        balances = loan.calculate_remaining_balances()
        outstanding = sum(balances.values())
        if outstanding <= 0:
            continue

        total_portfolio += outstanding
        total_loans_count += 1
        days = compute_installment_based_days_overdue(loan, today)["days_overdue"]

        for threshold, key in [
            (1, "par_1"),
            (30, "par_30"),
            (60, "par_60"),
            (90, "par_90"),
            (120, "par_120"),
            (180, "par_180"),
        ]:
            if days >= threshold:
                par[key] += outstanding

    pct = lambda v: (
        (v / total_portfolio * 100) if total_portfolio > 0 else Decimal("0.00")
    )

    return render(
        request,
        "loans/portfolio_at_risk_report.html",
        {
            "total_portfolio": total_portfolio,
            "total_loans": total_loans_count,
            "par": {
                "par_1_amount": par["par_1"],
                "par_1_pct": pct(par["par_1"]),
                "par_30_amount": par["par_30"],
                "par_30_pct": pct(par["par_30"]),
                "par_60_amount": par["par_60"],
                "par_60_pct": pct(par["par_60"]),
                "par_90_amount": par["par_90"],
                "par_90_pct": pct(par["par_90"]),
                "par_120_amount": par["par_120"],
                "par_120_pct": pct(par["par_120"]),
                "par_180_amount": par["par_180"],
                "par_180_pct": pct(par["par_180"]),
            },
            "report_date": today.strftime("%d %B %Y"),
            "title": "Portfolio at Risk (PAR) Summary Report",
        },
    )


@login_required
@admin_or_manager_or_staff_required
def non_performing_loans(request):
    today = timezone.now().date()
    loans_qs = (
        Loan.objects.filter(
            Q(status="overdue")
            | Q(due_date__lt=today, status__in=["disbursed", "approved"])
        )
        .select_related("borrower", "account")
        .prefetch_related("repayments")
    )

    loan_data = []
    for loan in loans_qs:
        balances = loan.calculate_remaining_balances()
        outstanding = sum(balances.values())
        if outstanding <= 0:
            continue

        last = loan.repayments.order_by("-repayment_date").first()
        days_overdue = (
            max((today - loan.due_date).days, 0)
            if loan.due_date and loan.due_date < today
            else 0
        )

        loan_data.append(
            {
                "loan_id": loan.id,
                "borrower": loan.borrower.full_name,
                "principal_amount": loan.principal_amount,
                "interest_rate": loan.interest_rate,
                "status": loan.status,
                "due_date": loan.due_date,
                "days_overdue": days_overdue,
                "outstanding_balance": outstanding,
                "last_payment": last.repayment_date if last else None,
                "next_payment": _get_next_payment_date(loan),
            }
        )

    loan_data.sort(key=lambda x: x["days_overdue"], reverse=True)
    return render(
        request,
        "loans/non_performing_loans.html",
        {
            "page_obj": paginate_queryset(loan_data, request.GET.get("page")),
            "table_title": "Non-Performing Loans with Outstanding Balance",
            "today": today,
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_due_overdue_report(request):
    try:
        timezone.activate(pytz.timezone("Africa/Nairobi"))
    except Exception:
        timezone.activate(pytz.UTC)

    selected_date_str = request.GET.get("selected_date")
    selected_date = (
        datetime.strptime(selected_date_str, "%Y-%m-%d").date()
        if selected_date_str
        else timezone.now().date()
    )

    search_term = request.GET.get("search", "").strip().lower()
    page_due = int(request.GET.get("page_due", 1))
    page_arrears = int(request.GET.get("page_arrears", 1))
    page_overdue = int(request.GET.get("page_overdue", 1))

    cache_key = f"due_overdue_{selected_date}_{search_term}_{page_due}_{page_arrears}_{page_overdue}"
    if not search_term:
        cached = cache.get(cache_key)
        if cached:
            return render(request, "loans/loan_overdue_report.html", cached)

    loans_qs = (
        Loan.objects.filter(status__in=["disbursed", "overdue"])
        .select_related("borrower")
        .prefetch_related("repayments", "penalties")
        .order_by("id")
    )

    due_today_list, arrears_list, past_maturity_list = [], [], []

    for loan in loans_qs:
        try:
            if not loan.disbursement_date or loan.loan_period_months <= 0:
                continue

            if search_term:
                borrower = loan.borrower
                searchable = " ".join(
                    [
                        str(loan.id),
                        getattr(borrower, "full_name", "").lower(),
                        getattr(borrower, "phone_number", "").lower(),
                        str(loan.principal_amount),
                    ]
                )
                if search_term not in searchable:
                    continue

            balances = loan.calculate_remaining_balances()
            outstanding = sum(balances.values())
            if outstanding <= 0:
                continue

            schedule = loan.generate_payment_schedule() or []
            payments = [
                {
                    **p,
                    "payment_due_date": (
                        p["payment_due_date"].date()
                        if isinstance(p["payment_due_date"], datetime)
                        else p["payment_due_date"]
                    ),
                }
                for p in schedule
                if (p.get("principal_payment", 0) + p.get("interest_payment", 0)) > 0
            ]

            base = {
                "loan": loan,
                "principal_balance": balances["principal_balance"],
                "interest_balance": balances["interest_balance"],
                "penalty_balance": balances["penalty_balance"],
                "total_outstanding": outstanding,
            }

            # DUE TODAY
            due_today = [p for p in payments if p["payment_due_date"] == selected_date]
            if due_today:
                expected = sum(
                    p["principal_payment"] + p["interest_payment"] for p in due_today
                )
                due_amount = loan.calculate_total_amount_due_balance(
                    selected_date, Decimal(str(expected))
                )
                if due_amount > 0:
                    due_today_list.append({**base, "due_amount": due_amount})
                continue

            # ARREARS (missed installments, maturity not yet passed)
            missed = [p for p in payments if p["payment_due_date"] < selected_date]
            if missed and (not loan.due_date or loan.due_date >= selected_date):
                earliest = min(p["payment_due_date"] for p in missed)
                days_arrears = (selected_date - earliest).days
                expected = sum(
                    p["principal_payment"] + p["interest_payment"] for p in missed
                )
                arrears_amount = loan.calculate_total_amount_due_balance(
                    selected_date, Decimal(str(expected))
                )
                if arrears_amount > 0:
                    arrears_list.append(
                        {
                            **base,
                            "arrears_amount": arrears_amount,
                            "days_arrears": days_arrears,
                        }
                    )

            # PAST MATURITY
            if loan.due_date and loan.due_date < selected_date:
                days_past = (selected_date - loan.due_date).days
                mat_amount = loan.calculate_total_amount_due_balance(
                    selected_date, outstanding
                )
                if mat_amount > 0:
                    past_maturity_list.append(
                        {
                            **base,
                            "maturity_amount": mat_amount,
                            "days_past_maturity": days_past,
                        }
                    )

        except Exception as e:
            logger.error("Loan %s error in due/overdue report: %s", loan.id, e)

    per_page = 50
    context = {
        "due_loans": Paginator(due_today_list, per_page).get_page(page_due),
        "arrears_loans": Paginator(arrears_list, per_page).get_page(page_arrears),
        "overdue_loans": Paginator(past_maturity_list, per_page).get_page(page_overdue),
        "due_loans_count": len(due_today_list),
        "arrears_loans_count": len(arrears_list),
        "overdue_loans_count": len(past_maturity_list),
        "due_loans_total": sum(i["total_outstanding"] for i in due_today_list),
        "arrears_loans_total": sum(i["total_outstanding"] for i in arrears_list),
        "overdue_loans_total": sum(i["total_outstanding"] for i in past_maturity_list),
        "selected_date": selected_date,
        "now": timezone.now(),
        "table_title": "Due, Arrears & Past Maturity Loans Report",
    }

    if not search_term:
        cache.set(cache_key, context, 3600)

    return render(request, "loans/loan_overdue_report.html", context)


@login_required
@admin_or_manager_or_staff_required
def client_loan_statement(request):
    clients = (
        Client.objects.filter(loans__status__in=["disbursed", "overdue", "repaid"])
        .distinct()
        .order_by("full_name")
    )
    client, statement_data = None, None

    if request.method == "POST":
        client_id = request.POST.get("client_id")
        if client_id:
            client = get_object_or_404(Client, id=client_id)
            statement_data = []

            for loan in (
                Loan.objects.filter(borrower=client)
                .select_related("account")
                .prefetch_related("repayments", "transactions", "penalties")
                .order_by("-created_at")
            ):
                repayments = loan.repayments.all().order_by("repayment_date")
                totals = repayments.aggregate(
                    total_principal=Sum("principal_payment"),
                    total_interest=Sum("interest_payment"),
                    total_penalty=Sum("penalty_payment"),
                )
                tp = totals["total_principal"] or Decimal("0.00")
                ti = totals["total_interest"] or Decimal("0.00")
                tk = totals["total_penalty"] or Decimal("0.00")

                total_penalties = loan.penalties.aggregate(t=Sum("penalty_amount"))[
                    "t"
                ] or Decimal("0.00")
                pb = loan.principal_amount - tp
                ib = (loan.total_interest or Decimal("0")) - ti
                kb = total_penalties - tk

                statement_data.append(
                    {
                        "loan": loan,
                        "repayments": repayments,
                        "transactions": loan.transactions.all().order_by(
                            "transaction_date"
                        ),
                        "principal_balance": pb,
                        "interest_balance": ib,
                        "penalty_balance": kb,
                        "total_balance": pb + ib + kb,
                        "payment_schedule": loan.generate_payment_schedule(),
                    }
                )

    return render(
        request,
        "loans/loan_statement.html",
        {
            "clients": clients,
            "client": client,
            "statement_data": statement_data,
        },
    )


@login_required
@admin_or_manager_or_staff_required
def loan_penalty_management(request):
    clients_with_loans = (
        Client.objects.filter(loans__isnull=False).distinct().order_by("full_name")
    )
    selected_client = None
    unpaid_penalties, paid_penalties = [], []
    unpaid_total = paid_total = total_ever = Decimal("0.00")
    unpaid_count = paid_count = 0

    if request.method == "POST":
        client_id = request.POST.get("client_id")
        if client_id:
            selected_client = get_object_or_404(Client, id=client_id)

            if "delete_selected" in request.POST:
                ids = request.POST.getlist("penalty_ids")
                if ids:
                    with transaction.atomic():
                        penalties = list(
                            LoanPenalty.objects.select_related(
                                "loan", "account"
                            ).filter(
                                id__in=ids,
                                loan__borrower=selected_client,
                                is_deleted=False,
                            )
                        )
                        paid_count_blocked = sum(
                            1 for p in penalties if p.is_paid or p.remaining_amount <= 0
                        )
                        reversible = [
                            p
                            for p in penalties
                            if not p.is_paid and p.remaining_amount > 0
                        ]
                        reversed_total = Decimal("0.00")
                        for penalty in reversible:
                            reversed_total += _reverse_penalty_balance(
                                penalty, request.user
                            )

                    if reversible:
                        messages.success(
                            request,
                            (
                                f"Reversed {len(reversible)} unpaid penalt"
                                f"{'y' if len(reversible) == 1 else 'ies'} totaling "
                                f"{reversed_total:,.2f} UGX."
                            ),
                            extra_tags="bg-success",
                        )
                    if paid_count_blocked:
                        messages.warning(
                            request,
                            f"{paid_count_blocked} paid or cleared penalt{'y was' if paid_count_blocked == 1 else 'ies were'} left unchanged.",
                            extra_tags="bg-warning",
                        )
                else:
                    messages.warning(
                        request,
                        "Select at least one unpaid penalty to reverse.",
                        extra_tags="bg-warning",
                    )

            unpaid_penalties = (
                LoanPenalty.objects.filter(
                    loan__borrower=selected_client,
                    is_paid=False,
                    remaining_amount__gt=0,
                    is_deleted=False,
                )
                .select_related("loan")
                .order_by("-penalty_date")
            )
            unpaid_total = unpaid_penalties.aggregate(t=Sum("remaining_amount"))[
                "t"
            ] or Decimal("0")
            unpaid_count = unpaid_penalties.count()

            # Build paid list from paid penalties linked to this client
            paid_penalty_ids = set()
            paid_penalties = []

            for repayment in (
                LoanRepayment.objects.filter(
                    loan__borrower=selected_client, penalty_payment__gt=0
                )
                .select_related("loan")
                .order_by("-repayment_date")
            ):
                for penalty in LoanPenalty.objects.filter(
                    loan=repayment.loan,
                    is_paid=True,
                    updated_at__gte=repayment.repayment_date - timedelta(minutes=5),
                    updated_at__lte=repayment.repayment_date + timedelta(minutes=5),
                    is_deleted=False,
                ):
                    if penalty.id not in paid_penalty_ids:
                        paid_penalty_ids.add(penalty.id)
                        paid_penalties.append(
                            {
                                "penalty": penalty,
                                "paid_on": repayment.repayment_date,
                                "paid_via_repayment": repayment.id,
                            }
                        )

            for penalty in LoanPenalty.objects.filter(
                loan__borrower=selected_client, is_paid=True, is_deleted=False
            ).exclude(id__in=paid_penalty_ids):
                paid_penalties.append(
                    {
                        "penalty": penalty,
                        "paid_on": penalty.updated_at.date(),
                        "paid_via_repayment": None,
                    }
                )

            paid_total = sum(p["penalty"].penalty_amount for p in paid_penalties)
            paid_count = len(paid_penalties)
            total_ever = unpaid_total + paid_total

    return render(
        request,
        "loans/loan_penalty_management.html",
        {
            "table_title": "Loan Penalties Management",
            "clients": clients_with_loans,
            "selected_client": selected_client,
            "unpaid_penalties": unpaid_penalties,
            "paid_penalties": paid_penalties,
            "unpaid_total": unpaid_total,
            "paid_total": paid_total,
            "total_ever": total_ever,
            "unpaid_count": unpaid_count,
            "paid_count": paid_count,
        },
    )


@login_required
def loan_reports_dashboard(request):
    return render(
        request,
        "loans/loan_reports.html",
        {
            "form_title": "Loan Management Dashboard",
        },
    )


STANDARD_LOAN_COLUMNS = [
    ReportColumn("loan_id", "Loan No."),
    ReportColumn("client", "Client"),
    ReportColumn("loan_product", "Product"),
    ReportColumn("loan_officer", "Loan Officer"),
    ReportColumn("status", "Status"),
    ReportColumn("disbursement_date", "Disbursed"),
    ReportColumn("maturity_date", "Maturity"),
    ReportColumn("principal", "Principal", "right", True),
    ReportColumn("interest", "Interest", "right", True),
    ReportColumn("paid_amount", "Paid", "right", True),
    ReportColumn("outstanding_principal", "Principal Bal.", "right", True),
    ReportColumn("outstanding_interest", "Interest Bal.", "right", True),
    ReportColumn("outstanding_penalties", "Penalties/Fees", "right", True),
    ReportColumn("outstanding_amount", "Outstanding", "right", True),
    ReportColumn("overdue_amount", "Overdue", "right", True),
    ReportColumn("days_in_arrears", "Days Arrears", "right"),
    ReportColumn("aging_bucket", "Aging Bucket"),
]

AGING_COLUMNS = [
    ReportColumn("client", "Client"),
    ReportColumn("loan_id", "Loan No."),
    ReportColumn("loan_product", "Product"),
    ReportColumn("loan_officer", "Officer"),
    ReportColumn("disbursement_date", "Disbursed"),
    ReportColumn("maturity_date", "Maturity"),
    ReportColumn("outstanding_principal", "Principal Bal.", "right", True),
    ReportColumn("outstanding_interest", "Interest Bal.", "right", True),
    ReportColumn("outstanding_penalties", "Penalties/Fees", "right", True),
    ReportColumn("outstanding_amount", "Total Outstanding", "right", True),
    ReportColumn("overdue_amount", "Overdue", "right", True),
    ReportColumn("days_in_arrears", "Days Arrears", "right"),
    ReportColumn("aging_bucket", "Aging Bucket"),
    ReportColumn("last_repayment_date", "Last Payment"),
]

COLLECTION_COLUMNS = [
    ReportColumn("repayment_date", "Date"),
    ReportColumn("loan_id", "Loan No."),
    ReportColumn("client", "Client"),
    ReportColumn("principal", "Principal", "right", True),
    ReportColumn("interest", "Interest", "right", True),
    ReportColumn("fees", "Fees", "right", True),
    ReportColumn("penalties", "Penalties", "right", True),
    ReportColumn("paid_amount", "Total Paid", "right", True),
    ReportColumn("account", "Account"),
    ReportColumn("description", "Description"),
]

PAR_COLUMNS = [
    ReportColumn("bucket", "Risk Band"),
    ReportColumn("loan_count", "Loans", "right"),
    ReportColumn("outstanding_amount", "Outstanding", "right", True),
    ReportColumn("portfolio_percent", "% of Portfolio", "right"),
]

DUE_OVERDUE_COLUMNS = [
    ReportColumn("category", "Category"),
    ReportColumn("client", "Client"),
    ReportColumn("loan_id", "Loan No."),
    ReportColumn("loan_product", "Product"),
    ReportColumn("loan_officer", "Officer"),
    ReportColumn("disbursement_date", "Disbursed"),
    ReportColumn("maturity_date", "Maturity"),
    ReportColumn("expected_due", "Expected Due", "right", True),
    ReportColumn("overdue_amount", "Due/Overdue", "right", True),
    ReportColumn("outstanding_amount", "Outstanding", "right", True),
    ReportColumn("days_in_arrears", "Days Arrears", "right"),
    ReportColumn("aging_bucket", "Aging Bucket"),
]

PERFORMANCE_COLUMNS = [
    ReportColumn("loan_officer", "Loan Officer"),
    ReportColumn("loan_count", "Loans", "right"),
    ReportColumn("principal", "Principal", "right", True),
    ReportColumn("paid_amount", "Collected", "right", True),
    ReportColumn("outstanding_amount", "Outstanding", "right", True),
    ReportColumn("overdue_amount", "Overdue", "right", True),
]

LOAN_TOTAL_KEYS = [
    "principal",
    "interest",
    "paid_amount",
    "outstanding_principal",
    "outstanding_interest",
    "outstanding_penalties",
    "outstanding_amount",
    "overdue_amount",
]

COLLECTION_TOTAL_KEYS = ["principal", "interest", "fees", "penalties", "paid_amount"]


def _loan_report_rows(
    filters, *, date_field="disbursement_date", statuses=None, as_of=None
):
    qs = filtered_loans(filters, date_field=date_field)
    if statuses:
        qs = qs.filter(status__in=statuses)
    return [loan_financial_row(loan, today=as_of) for loan in qs]


@login_required
@admin_or_manager_or_staff_required
def loan_aging_report(request):
    filters = parse_report_filters(request)
    rows = [
        row
        for row in _loan_report_rows(filters, statuses=["disbursed", "overdue"])
        if row["outstanding_amount"] > 0
    ]
    rows.sort(key=lambda item: (item["days_in_arrears"], item["client"]))
    return _standard_report_response(
        request,
        "Loan Aging Report",
        "loan_aging_report.csv",
        AGING_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
        group_by="aging_bucket",
    )


@login_required
@admin_or_manager_or_staff_required
def loan_arrears_report(request):
    filters = parse_report_filters(request)
    rows = [
        row
        for row in _loan_report_rows(filters, statuses=["disbursed", "overdue"])
        if row["days_in_arrears"] > 0 and row["outstanding_amount"] > 0
    ]
    rows.sort(key=lambda item: (-item["days_in_arrears"], -item["outstanding_amount"]))
    return _standard_report_response(
        request,
        "Loan Arrears Report",
        "loan_arrears_report.csv",
        AGING_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
        group_by="aging_bucket",
    )


@login_required
@admin_or_manager_or_staff_required
def loan_portfolio_report(request):
    filters = parse_report_filters(request)
    rows = [row for row in _loan_report_rows(filters) if row["outstanding_amount"] > 0]
    return _standard_report_response(
        request,
        "Loan Portfolio Summary",
        "loan_portfolio_summary.csv",
        STANDARD_LOAN_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def portfolio_at_risk(request):
    filters = parse_report_filters(request)
    portfolio_rows = [
        row
        for row in _loan_report_rows(filters, statuses=["disbursed", "overdue"])
        if row["outstanding_amount"] > 0
    ]
    par = portfolio_at_risk_summary(portfolio_rows)
    rows = [
        {
            **band,
            "portfolio_percent": f"{band['portfolio_percent']:.2f}%",
        }
        for band in par["bands"]
    ]
    return _standard_report_response(
        request,
        "Portfolio at Risk Report",
        "portfolio_at_risk_report.csv",
        PAR_COLUMNS,
        rows,
        [],
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def non_performing_loans(request):
    filters = parse_report_filters(request)
    rows = [
        row
        for row in _loan_report_rows(filters, statuses=["disbursed", "overdue"])
        if row["days_in_arrears"] >= 90 and row["outstanding_amount"] > 0
    ]
    rows.sort(key=lambda item: (-item["days_in_arrears"], -item["outstanding_amount"]))
    return _standard_report_response(
        request,
        "Non-Performing Loans Report",
        "non_performing_loans_report.csv",
        AGING_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
        group_by="aging_bucket",
    )


@login_required
@admin_or_manager_or_staff_required
def loan_due_overdue_report(request):
    filters = parse_report_filters(request)
    selected_date = parse_date(request.GET.get("date") or "") or timezone.localdate()
    rows = []
    for row in _loan_report_rows(
        filters, statuses=["disbursed", "overdue"], as_of=selected_date
    ):
        if row["outstanding_amount"] <= 0:
            continue
        if row["overdue_amount"] <= 0:
            continue
        if row["days_in_arrears"] > 0:
            category = "In arrears"
        else:
            category = "Due today"
        if row["maturity_date"] and row["maturity_date"] < selected_date:
            category = "Past maturity"
        rows.append({**row, "category": category})
    rows.sort(
        key=lambda item: (item["category"], -item["days_in_arrears"], item["client"])
    )
    return _standard_report_response(
        request,
        "Due, Arrears And Past Maturity Report",
        "due_arrears_past_maturity_report.csv",
        DUE_OVERDUE_COLUMNS,
        rows,
        ["expected_due", "overdue_amount", "outstanding_amount"],
        filters,
        group_by="category",
    )


@login_required
@admin_or_manager_or_staff_required
def loan_disbursement_report(request):
    filters = parse_report_filters(request)
    rows = [
        loan_financial_row(loan)
        for loan in filtered_loans(filters).filter(disbursement_date__isnull=False)
    ]
    return _standard_report_response(
        request,
        "Loan Disbursement Report",
        "loan_disbursement_report.csv",
        STANDARD_LOAN_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def loan_collection_report(request):
    filters = parse_report_filters(request)
    rows = repayment_rows(filters)
    return _standard_report_response(
        request,
        "Loan Collection Report",
        "loan_collection_report.csv",
        COLLECTION_COLUMNS,
        rows,
        COLLECTION_TOTAL_KEYS,
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def outstanding_loan_balances_report(request):
    filters = parse_report_filters(request)
    rows = [
        row
        for row in (loan_financial_row(loan) for loan in filtered_loans(filters))
        if row["outstanding_amount"] > 0
    ]
    return _standard_report_response(
        request,
        "Outstanding Loan Balances Report",
        "outstanding_loan_balances_report.csv",
        STANDARD_LOAN_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def defaulted_loans_report(request):
    filters = parse_report_filters(request)
    rows = [
        row
        for row in (loan_financial_row(loan) for loan in filtered_loans(filters))
        if row["days_in_arrears"] > 90 and row["outstanding_amount"] > 0
    ]
    return _standard_report_response(
        request,
        "Defaulted Loans Report",
        "defaulted_loans_report.csv",
        STANDARD_LOAN_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
        group_by="aging_bucket",
    )


@login_required
@admin_or_manager_or_staff_required
def closed_loans_report(request):
    filters = parse_report_filters(request)
    rows = [
        loan_financial_row(loan)
        for loan in filtered_loans(filters, date_field="updated_at").filter(
            status__in=["closed", "repaid"]
        )
    ]
    return _standard_report_response(
        request,
        "Closed Loans Report",
        "closed_loans_report.csv",
        STANDARD_LOAN_COLUMNS,
        rows,
        LOAN_TOTAL_KEYS,
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def loan_officer_performance_report(request):
    filters = parse_report_filters(request)
    rows = []
    grouped = {}
    for loan in filtered_loans(filters):
        row = loan_financial_row(loan)
        officer = row["loan_officer"]
        grouped.setdefault(
            officer,
            {
                "loan_officer": officer,
                "loan_count": 0,
                "principal": Decimal("0.00"),
                "paid_amount": Decimal("0.00"),
                "outstanding_amount": Decimal("0.00"),
                "overdue_amount": Decimal("0.00"),
            },
        )
        grouped[officer]["loan_count"] += 1
        for key in ["principal", "paid_amount", "outstanding_amount", "overdue_amount"]:
            grouped[officer][key] += row[key]
    rows = sorted(grouped.values(), key=lambda item: item["loan_officer"])
    return _standard_report_response(
        request,
        "Loan Officer Performance Report",
        "loan_officer_performance_report.csv",
        PERFORMANCE_COLUMNS,
        rows,
        ["principal", "paid_amount", "outstanding_amount", "overdue_amount"],
        filters,
    )


@login_required
@admin_or_manager_or_staff_required
def loan_product_performance_report(request):
    filters = parse_report_filters(request)
    grouped = {}
    for loan in filtered_loans(filters):
        row = loan_financial_row(loan)
        product = row["loan_product"]
        grouped.setdefault(
            product,
            {
                "loan_product": product,
                "loan_count": 0,
                "principal": Decimal("0.00"),
                "paid_amount": Decimal("0.00"),
                "outstanding_amount": Decimal("0.00"),
                "overdue_amount": Decimal("0.00"),
            },
        )
        grouped[product]["loan_count"] += 1
        for key in ["principal", "paid_amount", "outstanding_amount", "overdue_amount"]:
            grouped[product][key] += row[key]
    rows = sorted(grouped.values(), key=lambda item: item["loan_product"])
    columns = [
        ReportColumn("loan_product", "Loan Product"),
        ReportColumn("loan_count", "Loans", "right"),
        ReportColumn("principal", "Principal", "right", True),
        ReportColumn("paid_amount", "Collected", "right", True),
        ReportColumn("outstanding_amount", "Outstanding", "right", True),
        ReportColumn("overdue_amount", "Overdue", "right", True),
    ]
    return _standard_report_response(
        request,
        "Loan Product Performance Report",
        "loan_product_performance_report.csv",
        columns,
        rows,
        ["principal", "paid_amount", "outstanding_amount", "overdue_amount"],
        filters,
    )


def _standard_report_response(
    request,
    title,
    csv_filename,
    columns,
    rows,
    total_keys,
    filters,
    *,
    group_by=None,
):
    if filters.get("export") == "csv":
        return export_rows_csv(csv_filename, columns, rows)

    page_obj = paginate_rows(rows, request.GET.get("page"), filters.get("per_page", 50))
    grouped_rows = group_rows_by_bucket(rows, group_by, total_keys) if group_by else []
    filter_form = LoanReportFilterForm(request.GET or None)
    filter_form.is_valid()

    return render(
        request,
        "loans/standard_report.html",
        {
            "organization_name": "Pendeza Uganda",
            "table_title": title,
            "columns": columns,
            "rows": list(page_obj.object_list),
            "all_rows": rows,
            "grouped_rows": grouped_rows,
            "group_by": group_by,
            "page_obj": page_obj,
            "totals": summarize_amounts(rows, total_keys),
            "total_keys": set(total_keys),
            "filters": filters,
            "filter_form": filter_form,
            "generated_at": timezone.now(),
            "status_choices": Loan.STATUS_CHOICES,
            "loan_product_choices": Loan.LOAN_PURPOSE_CHOICES,
            "clients": Client.objects.order_by("full_name"),
            "loan_officers": User.objects.filter(applied_loans__isnull=False)
            .distinct()
            .order_by("username"),
            "csv_url": _csv_url(request),
        },
    )


def _csv_url(request):
    params = request.GET.copy()
    params["export"] = "csv"
    return f"{request.path}?{params.urlencode()}"
