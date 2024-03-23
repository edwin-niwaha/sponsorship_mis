# from formtools.wizard.views import SessionWizardView
from django.http import  HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from openpyxl import load_workbook
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


from .models import *
from .forms import *

def home(request):
    return render(request, "users/home.html")

# =================================== The dashboard ===================================
@login_required
def dashboard(request):
    c_records = Child.objects.all()
    total_records = c_records.count()

    context = {
        "kids_registered": total_records,
        "total_no_of_kids": total_records,
    }
    return render(request, "main/dashboard.html", context)

# =================================== Fetch and display all children details ===================================
def child_list(request):
    # queryset = Child.objects.all().order_by('-created_at')
    queryset = Child.objects.all().order_by('id').select_related('profile_picture')    

    search_query = request.GET.get('search')
    if search_query:
        queryset = queryset.filter(full_name__icontains=search_query)

    paginator = Paginator(queryset, 10)  # Show 10 records per page
    page = request.GET.get('page')

    try:
        c_records = paginator.page(page)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page.
        c_records = paginator.page(1)
    except EmptyPage:
        # If page is out of range (e.g. 9999), deliver last page of results.
        c_records = paginator.page(paginator.num_pages)

    return render(request, 'main/child/manage_child.html', {'c_records': c_records, 'table_title': 'Children MasterList'})

# =================================== Fetch and display selected child details ===================================
@login_required
def child_details(request, pk):
    record = Child.objects.get(pk=pk)
    age = record.calculate_age()
    
    context = {'table_title': 'Child Profile Report', 'record': record, 'age': age}
    return render(request, 'main/child/child_profile_rpt.html', context)

# =================================== Register Child ===================================
@login_required
@transaction.atomic
def register_child(request):
  if request.method == 'POST':
    form = ChildForm(request.POST, request.FILES)

    if form.is_valid(): 
      form.save()
      messages.info(request, "Record saved successfully!", extra_tags="bg-success")

    else:
      # Display form errors
      return render(request, 'main/child/child_frm.html', {'form': form})
  else:
    form = ChildForm()
  return render(request, 'main/child/child_frm.html', {'form_name':'Child Registration', 'form': form})

# =================================== Upload Profile Picture ===================================
@login_required
@transaction.atomic
def update_picture(request):
    if request.method == 'POST':
        form = ChildProfilePictureForm(request.POST, request.FILES)
        if form.is_valid():
            child_id = request.POST.get('id')
            try:
                child_profile = Child.objects.get(pk=child_id)
            except Child.DoesNotExist:
                messages.error(request, "Child with ID {} does not exist.".format(child_id))
                return redirect("update_picture")

            # child_picture = form.save(commit=False)  # Avoid unnecessary save
            # child_picture.child = child_profile  # Set the child relationship
            # child_picture.save()

            # Set the new picture as current and deactivate previous current picture
            new_picture = form.save(commit=False)
            new_picture.child = child_profile
            new_picture.is_current = True
            new_picture.save()

            messages.success(request, "Profile picture updated successfully!")
            return redirect("update_picture") 
        else:
            messages.error(request, "Form is invalid.")
    else:
        form = ChildProfilePictureForm()

    # Retrieve all child objects
        children = Child.objects.all().order_by('-created_at')

    return render(request, 'main/child/profile_picture.html', {'form': form, 'form_name': 'Upload Profile Picture', 'children': children})

# =================================== Update Child data ===================================
@login_required
@transaction.atomic
def update_child(request, pk, template_name="main/child/child_frm.html"):
    try:
        child_record = Child.objects.get(pk=pk)
    except Child.DoesNotExist:
        messages.error(request, "Child record not found!")
        return redirect("child_list")  # Or a relevant error page

    if request.method == "POST":
        form = ChildForm(request.POST, request.FILES, instance=child_record)
        if form.is_valid():
            form.save()

            messages.success(request, "Child record updated successfully!")
            return redirect("child_list")  # Replace with the appropriate redirect URL
    else:
        # Pre-populate the form with existing data
        form = ChildForm(instance=child_record)

    context = {'form_name': 'Child Registration', 'form': form}
    return render(request, template_name, context)

# =================================== Deleted selected child ===================================
@login_required
@transaction.atomic
def delete_child(request, pk):
    c_records = Child.objects.get(id=pk)
    c_records.delete()
    messages.info(request, "Record deleted successfully!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("child_list"))

# =================================== Process and Import Excel data ===================================
@login_required
@transaction.atomic
# debug and refactor this code
def import_data(request):
    if request.method == 'POST':
        form = UploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                # Call process_and_import_data function
                process_and_import_data(excel_file)
                messages.success(request, "Data imported successfully!", extra_tags="bg-success")
            except Exception as e:
                messages.error(request, f"Error importing data: {e}", extra_tags="bg-danger")  # Handle unexpected errors
            return redirect("data_list")  # Replace with your redirect URL
    else:
        form = UploadForm()
    return render(request, 'main/child/import_frm.html', {'form_name': 'Import Excel Data', 'form': form})


# Function to import Excel data
def process_and_import_data(excel_file):
    try:
        wb = load_workbook(excel_file)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            fname = row[0].value
            preferred_name = row[1].value
            residence = row[2].value
            tribe = row[3].value
            gender = row[4].value
            date_of_birth = row[5].value
            weight = row[6].value
            height = row[7].value
            # avatar = row[8].value
            c_interest = row[8].value
            is_child_in_school = row[9].value
            name_of_the_school = row[10].value
            education_level = row[11].value
            child_class = row[12].value
            best_subject = row[13].value
            is_sponsored = row[14].value
            sponsorship_type = row[15].value
            father_name = row[16].value
            is_father_alive = row[17].value
            father_description = row[18].value
            mother_name = row[19].value
            is_mother_alive = row[20].value
            mother_description = row[21].value
            guardian = row[22].value
            guardian_contact = row[23].value
            relationship_with_guardian = row[24].value
            siblings = row[25].value
            background_info = row[26].value
            health_status = row[27].value
            responsibility = row[28].value
            relationship_with_christ = row[29].value
            religion = row[30].value
            prayer_request = row[31].value
            year_enrolled = row[32].value
            is_departed = row[33].value
            staff_comment = row[34].value
            compiled_by = row[35].value
            if fname is not None:
                obj = Child.objects.create(full_name=fname, preferred_name=preferred_name, residence=residence,
                                                 tribe=tribe, gender=gender, date_of_birth=date_of_birth, weight=weight,
                                                 height=height, c_interest=c_interest, is_child_in_school=is_child_in_school,
                                                 name_of_the_school=name_of_the_school, education_level=education_level,
                                                 child_class=child_class, best_subject=best_subject, is_sponsored=is_sponsored,
                                                 sponsorship_type=sponsorship_type, father_name=father_name, is_father_alive=is_father_alive,
                                                 father_description=father_description, mother_name=mother_name, is_mother_alive=is_mother_alive,
                                                 mother_description=mother_description, guardian=guardian, guardian_contact=guardian_contact, 
                                                 relationship_with_guardian=relationship_with_guardian, siblings=siblings, background_info=background_info,
                                                 health_status=health_status, responsibility=responsibility, relationship_with_christ=relationship_with_christ,
                                                 religion=religion, prayer_request= prayer_request, year_enrolled=year_enrolled, is_departed=is_departed, staff_comment=staff_comment,
                                                 compiled_by=compiled_by)
                obj.save()
    except Exception as e:
        raise e  # Reraise the exception for better error handling at the view level


# =================================== Fetch and display imported data ===================================
@login_required
def import_details(request):
    records = Child.objects.all()
    return render(request, 'main/child/data_list.html', {'table_title': 'Imported Excel Data', 'records': records})

# =================================== Delete selected individual ===================================
@login_required
@transaction.atomic
def delete_excel_data(request, pk):
    record_imported = Child.objects.get(id=pk)
    record_imported.delete()
    messages.info(request, "Record deleted!", extra_tags="bg-danger")
    return HttpResponseRedirect(reverse("data_list"))

# =================================== Delete all records at once ===================================
@login_required
@transaction.atomic
def delete_confirmation(request):
    if request.method == 'POST':
        Child.objects.all().delete()
        messages.info(request, "All records deleted!", extra_tags="bg-danger")
        return HttpResponseRedirect(reverse("data_list"))
    
# =================================== More ===================================
